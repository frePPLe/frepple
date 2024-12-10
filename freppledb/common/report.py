#
# Copyright (C) 2007-2019 by frePPLe bv
#
# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE
# LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION
# OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION
# WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#

r"""
This module implements a generic view to presents lists and tables.

It provides the following functionality:
 - Pagination of the results.
 - Ability to filter on fields, using different operators.
 - Ability to sort on a field.
 - Export the results as a CSV file, ready for use in a spreadsheet.
 - Import CSV formatted data files.
 - Show time buckets to show data by time buckets.
   The time buckets and time boundaries can easily be updated.
"""

import codecs
import csv
from datetime import date, datetime, timedelta, time
from decimal import Decimal
import functools
from hashlib import sha1
import logging
import math
import operator
import json
import random
import re
from time import timezone, localtime, sleep
from io import StringIO, BytesIO
import urllib
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import NamedStyle, PatternFill
from dateutil.parser import parse
from openpyxl.comments import Comment as CellComment

from django.apps import apps
from django.core.cache import cache
from django.db import connection
from django.db.models import Model, Lookup
from django.db.models.expressions import RawSQL
from django.db.utils import DEFAULT_DB_ALIAS, load_backend, OperationalError
from django.contrib.auth.models import Group
from django.contrib.auth import get_permission_codename
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.admin.utils import unquote, quote
from django.core.exceptions import ValidationError
from django.core.management.color import no_style
from django.db import connections, transaction, models
from django.db.models.fields import CharField, AutoField, DateField, DateTimeField
from django.db.models.fields.related import RelatedField
from django.forms.models import modelform_factory
from django.http import HttpResponse, StreamingHttpResponse, HttpResponseNotFound
from django.http import Http404, HttpResponseNotAllowed, HttpResponseForbidden
from django.shortcuts import render
from django.utils import translation
from django.utils.dateparse import parse_duration, parse_time
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, force_str
from django.utils.html import escape
from django.utils.translation import gettext_lazy as _
from django.utils.formats import get_format, date_format
from django.utils.text import capfirst, get_text_list, format_lazy
from django.contrib.contenttypes.models import ContentType
from django.views.generic.base import View

from freppledb import __version__
from freppledb.boot import getAttributeFields
from freppledb.common.auth import MultiDBBackend
from freppledb.common.models import (
    User,
    Comment,
    Parameter,
    BucketDetail,
    Bucket,
    HierarchyModel,
    NotificationFactory,
)
from freppledb.common.dataload import parseExcelWorksheet, parseCSVdata
from freppledb.common.localization import parseLocalizedDate, parseLocalizedDateTime
from freppledb.common.utils import getStorageUsage


logger = logging.getLogger(__name__)


# A list of models with some special, administrative purpose.
# They should be excluded from bulk import, export and erasing actions.
EXCLUDE_FROM_BULK_OPERATIONS = (Group, User, Comment)


separatorpattern = re.compile(r"[\s\-_]+")


@CharField.register_lookup
class IsChildOfLookup(Lookup):
    lookup_name = "ico"

    def as_sql(self, compiler, connection):
        lhs, lhs_params = self.process_lhs(compiler, connection)
        rhs, rhs_params = self.process_rhs(compiler, connection)
        params = lhs_params + rhs_params

        # getting the class name from the lhs argument
        objectModel = (
            lhs.split(".")[1][1:-4]
            .replace("origin", "location")
            .replace("destination", "location")
        )

        # rebuilding hierarchy
        import importlib

        c = getattr(
            importlib.import_module("freppledb.input.models"), objectModel.capitalize()
        )
        if issubclass(c, HierarchyModel):
            c.rebuildHierarchy(connection.alias)

        return (
            """
        %s in
        (select child.name from %s owner
        inner join %s child on child.lft between owner.lft and owner.rght
        where owner.name = %s)
        """
            % (lhs, objectModel, objectModel, rhs),
            params,
        )


def create_connection(alias=DEFAULT_DB_ALIAS):
    if alias == DEFAULT_DB_ALIAS:
        connections.configure_settings({alias: settings.DATABASES[alias]})
    else:
        connections.configure_settings(
            {
                DEFAULT_DB_ALIAS: settings.DATABASES[DEFAULT_DB_ALIAS],
                alias: settings.DATABASES[alias],
            }
        )
    db = connections.databases[alias]
    backend = load_backend(db["ENGINE"])
    return backend.DatabaseWrapper(db, alias)


def matchesModelName(name, model):
    """
    Returns true if the first argument is a valid name for the model passed as second argument.
    The string must match either:
      - the model name
      - the verbose name
      - the pural verbose name
    The comparison is case insensitive and also ignores whitespace, dashes and underscores.
    The comparison tries to find a match using the current active language, as well as in English.
    """
    checkstring = re.sub(separatorpattern, "", name.lower())
    # Try with the localized model names
    if checkstring == re.sub(separatorpattern, "", model._meta.model_name.lower()):
        return True
    elif checkstring == re.sub(separatorpattern, "", model._meta.verbose_name.lower()):
        return True
    elif checkstring == re.sub(
        separatorpattern, "", model._meta.verbose_name_plural.lower()
    ):
        return True
    else:
        # Try with English model names
        with translation.override("en"):
            if checkstring == re.sub(
                separatorpattern, "", model._meta.model_name.lower()
            ):
                return True
            elif checkstring == re.sub(
                separatorpattern, "", model._meta.verbose_name.lower()
            ):
                return True
            elif checkstring == re.sub(
                separatorpattern, "", model._meta.verbose_name_plural.lower()
            ):
                return True
            else:
                return False


def getHorizon(request, future_only=False):
    # Pick up the last current date or current date if last current date doesn't exist
    current = getCurrentDate(request.database, lastplan=True)
    horizontype = request.GET.get("horizontype", request.user.horizontype)
    horizonunit = request.GET.get("horizonunit", request.user.horizonunit)
    try:
        horizonlength = int(request.GET.get("horizonlength"))
    except Exception:
        horizonlength = request.user.horizonlength
    try:
        horizonbefore = int(request.GET.get("horizonbefore"))
    except Exception:
        horizonbefore = request.user.horizonbefore
    if horizontype:
        # First type: Horizon relative to the current date
        start = current.replace(hour=0, minute=0, second=0, microsecond=0)
        if horizonunit == "day":
            end = start + timedelta(days=horizonlength or 60)
            end = end.replace(hour=0, minute=0, second=0)
            if not future_only and horizonbefore:
                start -= timedelta(days=horizonbefore)
        elif horizonunit == "week":
            end = start + timedelta(weeks=horizonlength or 8, days=7 - start.weekday())
            if not future_only and horizonbefore:
                start -= timedelta(weeks=horizonbefore, days=start.weekday())
        else:
            y = start.year
            m = start.month + (horizonlength or 2) + (start.day > 1 and 1 or 0)
            while m > 12:
                y += 1
                m -= 12
            end = datetime(y, m, 1)
            y = start.year
            if not future_only and horizonbefore:
                m = start.month - horizonbefore + 1
                while m <= 0:
                    y -= 1
                    m += 12
                start = datetime(y, m, 1)
    else:
        # Second type: Absolute start and end dates given
        try:
            horizonstart = parseLocalizedDate(request.GET.get("horizonstart"))
        except Exception:
            horizonstart = request.user.horizonstart
        try:
            horizonend = parseLocalizedDate(request.GET.get("horizonend"))
        except Exception:
            horizonend = request.user.horizonend
        start = horizonstart
        if not start or (future_only and start < current):
            start = current.replace(hour=0, minute=0, second=0, microsecond=0)
        end = horizonend
        if end:
            if end < start:
                if future_only and end < current:
                    # Special case to assure a minimum number of future buckets
                    if horizonunit == "day":
                        end = start + timedelta(days=horizonlength or 60)
                    elif horizonunit == "week":
                        end = start + timedelta(weeks=horizonlength or 8)
                    else:
                        end = start + timedelta(weeks=horizonlength or 8)
                else:
                    # Swap start and end to assure the start is before the end
                    tmp = start
                    start = end
                    end = tmp
        else:
            if horizonunit == "day":
                end = start + timedelta(days=horizonlength or 60)
            elif horizonunit == "week":
                end = start + timedelta(weeks=horizonlength or 8)
            else:
                end = start + timedelta(weeks=horizonlength or 8)
    return (current, start, end)


class GridField:
    """
    Base field for columns in grid views.
    """

    def __init__(self, name, **kwargs):
        self.name = name
        for key, value in kwargs.items():
            setattr(self, key, value)
        if "key" in kwargs:
            self.editable = False
        if "title" not in kwargs and not self.title:
            self.title = self.name and _(self.name) or ""
        if not self.name:
            self.sortable = False
            self.search = False
        if "field_name" not in kwargs:
            self.field_name = self.name

    def __str__(self):
        o = [
            '"name":"%s","index":"%s","editable":%s,"label":"%s","align":"%s","title":false,"field_name":"%s"'
            % (
                self.name or "",
                self.name or "",
                self.editable and "true" or "false",
                force_str(self.title).title().replace("'", "\\'"),
                self.align,
                self.field_name,
            )
        ]
        if self.key:
            o.append(',"key":true')
        if not self.sortable:
            o.append(',"sortable":false')
        if not self.search:
            o.append(',"search":false')
        if self.formatter:
            o.append(',"formatter":"%s"' % self.formatter)
        if self.unformat:
            o.append(',"unformat":"%s"' % self.unformat)
        if self.searchrules:
            o.append(',"searchrules":{%s}' % self.searchrules)
        if self.hidden:
            o.append(',"alwayshidden":true, "hidden":true')
        if self.searchoptions:
            o.append(',"searchoptions":%s' % self.searchoptions)
        if self.extra:
            if callable(self.extra):
                o.append(",%s" % force_str(self.extra()))
            else:
                o.append(",%s" % force_str(self.extra))
        return "".join(o)

    name = None
    field_name = None
    formatter = None
    width = 100
    editable = True
    sortable = True
    search = True
    key = False
    unformat = None
    title = None
    extra = None
    align = "center"
    searchrules = None
    hidden = False  # NEVER display this field
    initially_hidden = False  # Hide the field by default, but allow the user to add it
    searchoptions = '{"searchhidden": true}'
    background_header = None  # Used when exporting to excel
    background_cell = None  # Used when exporting to excel


class GridFieldDateTime(GridField):
    formatter = "date"
    extra = (
        '"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"%s"}'
        % settings.DATETIME_FORMAT
    )
    searchoptions = (
        '{"sopt":["eq","ne","lt","le","gt","ge","win","isnull"],"searchhidden": true}'
    )
    width = 140


class GridFieldTime(GridField):
    formatter = "time"
    extra = '"formatoptions":{"srcformat":"H:i:s","newformat":"H:i:s"}'
    width = 80


class GridFieldDate(GridField):
    formatter = "date"
    extra = (
        '"formatoptions":{"srcformat":"Y-m-d","newformat":"%s"}' % settings.DATE_FORMAT
    )
    searchoptions = (
        '{"sopt":["eq","ne","lt","le","gt","ge","win","isnull"],"searchhidden":true}'
    )
    width = 140


class GridFieldInteger(GridField):
    formatter = "integer"
    extra = '"formatoptions":{"defaultValue": ""}'
    searchoptions = (
        '{sopt:["eq","ne","in","ni","lt","le","gt","ge","isnull"],"searchhidden":true}'
    )
    width = 70
    searchrules = '"integer":true'


class GridFieldNumber(GridField):
    formatter = "number"
    extra = '"formatoptions":{"defaultValue":"","decimalPlaces":"auto"}'
    searchoptions = (
        '{sopt:["eq","ne","in","ni","lt","le","gt","ge","isnull"],"searchhidden":true}'
    )
    width = 70
    searchrules = '"number":true'


class GridFieldBool(GridField):
    extra = '"formatoptions":{"disabled":false}, "edittype":"checkbox", "editoptions":{"value":"True:False"}'
    width = 60


class GridFieldLastModified(GridField):
    formatter = "date"
    extra = (
        '"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"%s H:i:s"}'
        % settings.DATE_FORMAT
    )
    searchoptions = '{"sopt":["eq","ne","lt","le","gt","ge","win"],"searchhidden":true}'
    title = _("last modified")
    editable = False
    width = 140


class GridFieldJSON(GridField):
    width = 200
    align = "left"
    searchoptions = '{"sopt":["cn","nc"],"searchhidden":true}'


class GridFieldLocalDateTime(GridFieldDateTime):
    extra = (
        '"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"%s H:i:s"}'
        % settings.DATE_FORMAT
    )
    pass


class GridFieldText(GridField):
    width = 200
    align = "left"
    searchoptions = '{"sopt":["cn","nc","eq","ne","lt","le","gt","ge","bw","bn","in","ni","ew","en","isnull"],"searchhidden":true}'


class GridFieldHierarchicalText(GridFieldText):
    searchoptions = '{"sopt":["cn","nc","eq","ne","lt","le","gt","ge","bw","bn","in","ni","ew","en","ico","isnull"],"searchhidden":true}'


class GridFieldChoice(GridField):
    width = 100
    align = "center"
    searchoptions = '{"sopt":["cn","nc","eq","ne","bw","bn","in","ni","ew","en"],"searchhidden":true}'

    def __init__(self, name, **kwargs):
        super().__init__(name, **kwargs)
        e = ['"formatter":"select", "edittype":"select", "editoptions":{"value":"']
        first = True
        self.choices = kwargs.get("choices", [])
        for i in self.choices:
            if first:
                first = False
                e.append("%s:" % i[0])
            else:
                e.append(";%s:" % i[0])
            e.append(i[1])
        e.append('"}')
        self.extra = format_lazy("{}" * len(e), *e)


class GridFieldBoolNullable(GridFieldChoice):
    width = 60

    def __init__(self, name, **kwargs):
        kwargs["choices"] = (("", ""), ("False", _("No")), ("True", _("Yes")))
        super().__init__(name, **kwargs)


def getCurrency():
    try:
        cur = Parameter.getValue("currency").split(",")
        if len(cur) < 2:
            return ("", " %s" % escape(cur[0]))
        else:
            return ("%s " % escape(cur[0]), " %s" % escape(cur[1]))
    except Exception:
        return ("", " $")


def getCurrentDate(database=DEFAULT_DB_ALIAS, lastplan=False):
    if lastplan:
        try:
            return parse(
                Parameter.getValue("last_currentdate", default=None, database=database)
            )
        except Exception:
            pass
    val = Parameter.getValue("currentdate", default=None, database=database)
    try:
        return parse(val)
    except Exception:
        n = datetime.now()
        return (
            datetime(n.year, n.month, n.day)
            if val and val.lower() == "today"
            else n.replace(microsecond=0)
        )


class GridFieldCurrency(GridField):
    formatter = "currencyWithBlanks"
    searchoptions = (
        '{sopt:["eq","ne","in","ni","lt","le","gt","ge","isnull"],"searchhidden":true}'
    )

    def extra(self):
        cur = getCurrency()
        return '"formatoptions":%s' % json.dumps(
            {"prefix": cur[0], "suffix": cur[1], "defaultvalue": ""}
        )

    width = 80


class GridFieldDuration(GridField):
    formatter = "duration"
    width = 80
    searchoptions = (
        '{sopt:["eq","ne","in","ni","lt","le","gt","ge"],"searchhidden":true}'
    )


class EncodedCSVReader:
    """
    A CSV reader which will iterate over lines in the CSV data buffer.
    The reader will scan the BOM header in the data to detect the right encoding.
    """

    def __init__(self, datafile, **kwds):
        # Read the file into memory
        # TODO Huge file uploads can overwhelm your system!
        data = datafile.read()
        # Detect the encoding of the data by scanning the BOM.
        # Skip the BOM header if it is found.
        if data.startswith(codecs.BOM_UTF32_BE):
            self.reader = StringIO(data.decode("utf_32_be"))
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF32_LE):
            self.reader = StringIO(data.decode("utf_32_le"))
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_BE):
            self.reader = StringIO(data.decode("utf_16_be"))
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_LE):
            self.reader = StringIO(data.decode("utf_16_le"))
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF8):
            self.reader = StringIO(data.decode("utf_8"))
            self.reader.read(1)
        else:
            # No BOM header found. We assume the data is encoded in the default CSV character set.
            self.reader = StringIO(data.decode(settings.CSV_CHARSET))
        self.csvreader = csv.reader(self.reader, **kwds)

    def __next__(self):
        return next(self.csvreader)

    def __iter__(self):
        return self


class GridReport(View):
    """
    The base class for all jqgrid views.
    The parameter values defined here are used as defaults for all reports, but
    can be overwritten.
    """

    # Points to template to be used
    template = "admin/base_site_grid.html"

    # The title of the report. Used for the window title
    title = ""

    # A optional text shown after the title in the content.
    # It is however not added in the page title or the breadcrumb name
    post_title = ""

    # Link to the documentation
    help_url = None

    # The resultset that returns a list of entities that are to be
    # included in the report.
    # This query is used to return the number of records.
    # It is also used to generate the actual results, in case no method
    # "query" is provided on the class.
    basequeryset = None

    # Specifies which column is used for an initial ordering
    default_sort = (0, "asc")

    # A model class from which we can inherit information.
    model = None

    # Allow editing in this report or not
    editable = True

    # Allow filtering of the results or not
    filterable = True

    # Include time bucket support in the report
    hasTimeBuckets = False
    hasTimeOnly = False

    # Allow to exclude time buckets in the past
    showOnlyFutureTimeBuckets = False

    # Default logic: if there is an argument to the report, we always show table + graph
    # New logic: if there is an argument, we can still choose whether or not to use table and/or graph
    # Not very clean, but doing otherwise is backward incompatible and needs changing quite some templates :-(
    new_arg_logic = False

    # Allow this report to automatically restore the previous filter
    # (unless a filter is already applied explicitly in the URL of course)
    autofilter = True

    # Hack to allow variable height depending on the detail position
    variableheight = True

    # Specify a minimum level for the time buckets available in the report.
    # Higher values (ie more granular) buckets can then not be selected.
    maxBucketLevel = None
    minBucketLevel = None

    # Show a select box in front to allow selection of records
    multiselect = True

    # Control the height of the grid. By default the full browser window is used.
    height = None

    # Number of columns frozen in the report
    frozenColumns = 0

    # A list with required user permissions to view the report
    permissions = ()

    # This text is displayed in the grid when there are no data to display.
    # It is intended to display a short onboarding message.
    message_when_empty = None

    # Defines the difference between height of the grid and its boundaries
    heightmargin = 85

    # Define a list of actions
    actions = None

    # display the duplication icon
    canDuplicate = True

    _attributes_added = False

    @classmethod
    def getKey(cls, request, *args, **kwargs):
        return "%s.%s" % (cls.__module__, cls.__name__)

    @classmethod
    def _localize(cls, value, decimal_separator, excel_duration_in_days=False):
        if callable(value):
            value = value()
        if isinstance(value, numericTypes):
            return (
                decimal_separator == "," and str(value).replace(".", ",") or str(value)
            )
        elif isinstance(value, datetime):
            return date_format(value, format="DATETIME_FORMAT", use_l10n=False)
        elif isinstance(value, date):
            return date_format(value, format="DATE_FORMAT", use_l10n=False)
        elif isinstance(value, timedelta):
            return _parseSeconds(value, excel_duration_in_days)
        elif isinstance(value, (list, tuple)):
            return "|".join([str(cls._localize(i, decimal_separator)) for i in value])
        else:
            return value

    @staticmethod
    def getBOM(encoding):
        try:
            # Get the official name of the encoding (since encodings can have many alias names)
            name = codecs.lookup(encoding).name
        except Exception:
            return ""  # Unknown encoding, without BOM header
        if name == "utf-32-be":
            return codecs.BOM_UTF32_BE
        elif name == "utf-32-le":
            return codecs.BOM_UTF32_LE
        elif name == "utf-16-be":
            return codecs.BOM_UTF16_BE
        elif name == "utf-16-le":
            return codecs.BOM_UTF16_LE
        elif name == "utf-8":
            return codecs.BOM_UTF8
        else:
            return ""

    @classmethod
    def getAppLabel(cls):
        """
        Return the name of the Django application which defines this report.
        """
        if hasattr(cls, "app_label"):
            return cls.app_label
        s = cls.__module__.split(".")
        for i in range(len(s), 0, -1):
            x = ".".join(s[0:i])
            if x in settings.INSTALLED_APPS:
                cls.app_label = s[i - 1]
                return cls.app_label
        raise Exception("Can't identify app of reportclass %s" % cls)

    # Extra variables added to the report template
    @classmethod
    def extra_context(cls, request, *args, **kwargs):
        return {}

    @staticmethod
    def _getJSONValue(data, field=None, request=None):
        if isinstance(data, str) or isinstance(data, (list, tuple)):
            return json.dumps(data)
        elif isinstance(data, timedelta):
            return data.total_seconds()
        elif data is None:
            return '""'
        elif (
            isinstance(data, datetime)
            and isinstance(field, (GridFieldLastModified, GridFieldLocalDateTime))
            and request
        ):
            if not hasattr(request, "tzoffset"):
                request.tzoffset = GridReport.getTimezoneOffset(request)
            return '"%s"' % (data + request.tzoffset)
        else:
            return '"%s"' % data

    @classmethod
    def _getCSVValue(
        cls,
        data,
        field=None,
        request=None,
        decimal_separator="",
        excel_duration_in_days=False,
    ):
        if data is None:
            return ""
        else:
            if (
                isinstance(data, datetime)
                and isinstance(field, (GridFieldLastModified, GridFieldLocalDateTime))
                and request
            ):
                if not hasattr(request, "tzoffset"):
                    request.tzoffset = GridReport.getTimezoneOffset(request)
                data += request.tzoffset
            return force_str(
                cls._localize(data, decimal_separator, excel_duration_in_days),
                encoding=settings.CSV_CHARSET,
                errors="ignore",
            )

    @classmethod
    def getBuckets(cls, request, *args, **kwargs):
        """
        This function gets passed a name of a bucketization.
        It returns a tuple with:
          - the start date of the report horizon
          - the end date of the reporting horizon
          - a list of buckets.

        The functions takes into consideration some special flags:
          - showOnlyFutureTimeBuckets: filter to allow only future time buckets to be shown
          - maxBucketLevel: respect the lowest supported level in the time bucket hierarchy
          - minBucketLevel: respect the highest supported level in the time bucket hierarchy
        """
        # Select the bucket size
        if not cls.maxBucketLevel:
            maxlvl = 999
        elif callable(cls.maxBucketLevel):
            maxlvl = cls.maxBucketLevel(request)
        else:
            maxlvl = cls.maxBucketLevel
        if not cls.minBucketLevel:
            minlvl = -999
        elif callable(cls.minBucketLevel):
            minlvl = cls.minBucketLevel(request)
        else:
            minlvl = cls.minBucketLevel
        arg_buckets = request.GET.get("buckets", None)
        try:
            bucket = (
                Bucket.objects.using(request.database)
                .get(
                    name=arg_buckets or request.user.horizonbuckets,
                    level__lte=maxlvl,
                    level__gte=minlvl,
                )
                .name
            )
        except Exception:
            try:
                bucket = (
                    Bucket.objects.using(request.database)
                    .filter(level__lte=maxlvl, level__gte=minlvl)
                    .order_by("-level")[0]
                    .name
                )
            except Exception:
                bucket = None
        if not arg_buckets and not request.user.horizonbuckets and bucket:
            request.user.horizonbuckets = bucket
            request.user.save()

        # Get the report horizon
        current, start, end = getHorizon(
            request, future_only=cls.showOnlyFutureTimeBuckets
        )

        # Filter based on the start and end date
        request.current_date = str(current)
        request.report_startdate = start
        request.report_enddate = end
        request.report_bucket = str(bucket)
        if bucket and not getattr(cls, "hasTimeOnly", False):
            res = BucketDetail.objects.using(request.database).filter(bucket=bucket)
            if start:
                res = res.filter(enddate__gt=start)
            if end:
                res = res.filter(startdate__lt=end)
            request.report_bucketlist = res.annotate(
                history=RawSQL("case when enddate < %s then 1 else 0 end", [current])
            ).values("name", "startdate", "enddate", "history")
        else:
            request.report_bucketlist = []

    @staticmethod
    def getTimezoneOffset(request):
        """
        Return the difference between the end user's UTC offset and the server's UTC offset
        """
        try:
            offset = int(request.COOKIES.get("tzoffset", 0))
        except Exception:
            offset = 0

        daylight = max(0, localtime().tm_isdst)
        return timedelta(seconds=timezone - offset - daylight * 3600)

    @classmethod
    def has_permission(cls, user):
        for perm in cls.permissions:
            if not user.has_perm("auth.%s" % perm[0]):
                return False
        if cls.model and "view" in cls.model._meta.default_permissions:
            return user.has_perm(
                "%s.view_%s" % (cls.model._meta.app_label, cls.model._meta.model_name)
            )
        return True

    @method_decorator(staff_member_required)
    @method_decorator(csrf_protect)
    def dispatch(self, request, *args, **kwargs):
        # Verify the user is authorized to view the report
        if not self.has_permission(request.user):
            return HttpResponseForbidden("<h1>%s</h1>" % _("Permission denied"))

        # Unescape special characters in the arguments.
        # All arguments are encoded with escaping function used on the django admin.
        args_unquoted = [unquote(i) for i in args]

        # Add attributes if not done already
        if hasattr(self.__class__, "initialize"):
            self.__class__.initialize(request)
        if not self._attributes_added and self.model:
            self.__class__._attributes_added = True
            for f in getAttributeFields(self.model):
                self.__class__.rows += (f,)

        # Set row and cross attributes on the request
        if hasattr(self, "rows"):
            if callable(self.rows):
                request.rows = self.rows(request, *args, **kwargs)
            else:
                request.rows = self.rows
        if hasattr(self, "crosses"):
            if callable(self.crosses):
                request.crosses = self.crosses(request, *args, **kwargs)
            else:
                request.crosses = self.crosses

        # Dispatch to the correct method
        if request.method == "GET":
            return self.get(request, *args_unquoted, **kwargs)
        elif request.method == "POST":
            return self.post(request, *args_unquoted, **kwargs)
        else:
            return HttpResponseNotAllowed(["get", "post"])

    @classmethod
    def _validate_rows(cls, request, prefs):
        if not prefs:
            return [
                (
                    i,
                    request.rows[i].hidden or request.rows[i].initially_hidden,
                    request.rows[i].width,
                )
                for i in range(len(request.rows))
            ]
        else:
            # Validate the preferences to 1) map from name to index, 2) assure all rows
            # are included, 3) ignore non-existing fields
            defaultrows = {request.rows[i].name: i for i in range(len(request.rows))}
            rows = []
            for r in prefs:
                try:
                    idx = int(r[0])
                    if idx < len(request.rows):
                        defaultrows.pop(request.rows[idx].name, None)
                        rows.append(r)
                except (ValueError, IndexError):
                    if r[0] in defaultrows:
                        rows.append((defaultrows[r[0]], r[1], r[2]))
                        defaultrows.pop(r[0], None)
            for r, idx in defaultrows.items():
                rows.append(
                    (
                        idx,
                        request.rows[idx].hidden or request.rows[idx].initially_hidden,
                        request.rows[idx].width,
                    )
                )
            return rows

    @classmethod
    def _render_colmodel(
        cls, request, is_popup=False, prefs=None, mode="graph", *args, **kwargs
    ):
        if not prefs:
            frozencolumns = cls.frozenColumns
            rows = [
                (i, request.rows[i].initially_hidden, request.rows[i].width)
                for i in range(len(request.rows))
            ]
        else:
            frozencolumns = prefs.get("frozen", cls.frozenColumns)
            rows = cls._validate_rows(request, prefs.get("rows"))
        result = []
        if is_popup:
            result.append(
                '{"name":"select","label":gettext("Select"),"width":75,"align":"center","formatter":"selectbutton","sortable":false,"search":false}'
            )
        count = -1
        for index, hidden, width in rows:
            count += 1
            try:
                result.append(
                    '{%s,"width":%s,"counter":%d%s%s%s}'
                    % (
                        request.rows[index],
                        width,
                        index,
                        count < frozencolumns and ',"frozen":true' or "",
                        is_popup and ',"popup":true' or "",
                        hidden
                        and not request.rows[index].hidden
                        and ',"hidden":true'
                        or "",
                    )
                )
            except IndexError:
                logger.warning(
                    "Invalid preference value for %s: %s"
                    % (cls.getKey(request, *args, **kwargs), prefs)
                )
        return ",\n".join(result)

    @classmethod
    def _sanitize_excel_tab(cls, t):
        # Excel worksheet names are limited to 31 characters.
        # And they can't contain the characters \ * ? : / [ ]
        return (
            force_str(t)
            .replace(":", "")
            .replace("\\", "")
            .replace("?", "")
            .replace("[", "")
            .replace("]", "")
            .replace("/", "")[:31]
        )

    @classmethod
    def _generate_spreadsheet_data(
        cls, request, scenario_list, output, *args, **kwargs
    ):
        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue(
                "excel_duration_in_days", request.database, "false"
            ).lower()
            == "true"
        )

        # Create a workbook
        wb = Workbook(write_only=True)
        if callable(cls.title):
            title = cls.title(request, *args, **kwargs)
        else:
            title = cls.model._meta.verbose_name_plural if cls.model else cls.title
        wb.properties.creator = "frepple %s" % __version__
        ws = wb.create_sheet(title=cls._sanitize_excel_tab(title))

        # Create a named style for the header row
        headerstyle = NamedStyle(name="headerstyle")
        headerstyle.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        wb.add_named_style(headerstyle)
        readlonlyheaderstyle = NamedStyle(name="readlonlyheaderstyle")
        readlonlyheaderstyle.fill = PatternFill(fill_type="solid", fgColor="d0ebfb")
        wb.add_named_style(readlonlyheaderstyle)

        # Create custom named styles
        backgrounds = set()
        for r in request.rows:
            if r.background_header:
                backgrounds.add(r.background_header)
            if r.background_cell:
                backgrounds.add(r.background_cell)
        for bg in backgrounds:
            style = NamedStyle(name=bg)
            style.fill = PatternFill(fill_type="solid", fgColor=bg)
            wb.add_named_style(style)

        # Choose fields to export and write the title row

        # selected_rows comes from datasource url
        # rows can only be exported in English
        rows = request.GET.get("selected_rows", None)
        if rows:
            rows = rows.split(",")
            fields = [f for f in request.rows if f.name in rows]
            fields.sort(key=lambda x: rows.index(x.name))
        else:
            if not hasattr(request, "prefs"):
                request.prefs = request.user.getPreference(
                    cls.getKey(request, *args, **kwargs), database=request.database
                )
            if request.prefs and request.prefs.get("rows", None):
                # Customized settings
                fields = [
                    request.rows[f[0]]
                    for f in cls._validate_rows(request, request.prefs["rows"])
                    if not f[1] and not request.rows[f[0]].hidden
                ]
            else:
                # Default settings
                fields = [
                    i
                    for i in request.rows
                    if i.field_name and not i.hidden and not i.initially_hidden
                ]
        # Write a formatted header row
        header = []
        comment = None
        for f in fields:
            cell = WriteOnlyCell(ws, value=force_str(f.title).title())
            if f.editable or f.key:
                cell.style = f.background_header or "headerstyle"
                fname = getattr(f, "field_name", f.name)
                if not f.key and f.formatter == "detail" and fname.endswith("__name"):
                    cell.comment = CellComment(
                        force_str(
                            _("Values in this field must exist in the %s table")
                            % force_str(_(fname[:-6]))
                        ),
                        "Author",
                    )
                elif isinstance(f, GridFieldChoice):
                    cell.comment = CellComment(
                        force_str(
                            _("Accepted values are: %s")
                            % ", ".join([c[0] for c in f.choices])
                        ),
                        "Author",
                    )
            else:
                cell.style = f.background_header or "readlonlyheaderstyle"
                if not comment:
                    comment = CellComment(
                        force_str(_("Read only")), "Author", height=20, width=80
                    )
                cell.comment = comment
            header.append(cell)
        if len(scenario_list) > 1:
            cell = WriteOnlyCell(ws, value=force_str(_("scenario")).title())
            cell.style = "readlonlyheaderstyle"
            header.insert(0, cell)
        ws.append(header)

        # Add an auto-filter to the table
        ws.auto_filter.ref = "A1:%s1048576" % get_column_letter(len(header))

        original_database = request.database
        try:
            for scenario in scenario_list:
                request.database = scenario
                if hasattr(request, "query"):
                    delattr(request, "query")

                # Loop over all records
                for row in cls.data_query(request, *args, fields=fields, **kwargs):
                    r = []
                    if hasattr(row, "__getitem__"):
                        for f in fields:
                            cell = WriteOnlyCell(
                                ws,
                                value=_getCellValue(
                                    row[f.field_name],
                                    field=f,
                                    request=request,
                                    excel_duration_in_days=excel_duration_in_days,
                                ),
                            )
                            if f.background_cell:
                                cell.style = f.background_cell
                            r.append(cell)
                    else:
                        for f in fields:
                            cell = WriteOnlyCell(
                                ws,
                                value=_getCellValue(
                                    getattr(row, f.field_name),
                                    field=f,
                                    request=request,
                                    excel_duration_in_days=excel_duration_in_days,
                                ),
                            )
                            if f.background_cell:
                                cell.style = f.background_cell
                            r.append(cell)
                    if len(scenario_list) > 1:
                        r.insert(0, scenario_list[scenario])
                    ws.append(r)
        finally:
            request.database = original_database

        # Write the spreadsheet
        wb.save(output)

    @classmethod
    def _generate_csv_data(cls, request, scenario_list, *args, **kwargs):
        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue(
                "excel_duration_in_days", request.database, "false"
            ).lower()
            == "true"
        )

        sf = StringIO()
        decimal_separator = get_format("DECIMAL_SEPARATOR", request.LANGUAGE_CODE, True)
        if decimal_separator == ",":
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=";")
        else:
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=",")
        if translation.get_language() != request.LANGUAGE_CODE:
            translation.activate(request.LANGUAGE_CODE)

        # Write a Unicode Byte Order Mark header, aka BOM (Excel needs it to open UTF-8 file properly)
        yield cls.getBOM(settings.CSV_CHARSET)

        # Choose fields to export
        if not hasattr(request, "prefs"):
            request.prefs = request.user.getPreference(
                cls.getKey(request, *args, **kwargs), database=request.database
            )

        allColumns = request.GET.get("allcolumns", False)

        if allColumns and cls and request.rows:
            r = [
                force_str(
                    f.title, encoding=settings.CSV_CHARSET, errors="ignore"
                ).title()
                for f in request.rows
            ]
            if len(scenario_list) > 1:
                r.insert(0, _("scenario"))
            writer.writerow(r)
            fields = [f for f in request.rows]
        elif request.prefs and request.prefs.get("rows", None):
            # Customized settings
            custrows = cls._validate_rows(request, request.prefs["rows"])
            r = [
                force_str(
                    request.rows[f[0]].title,
                    encoding=settings.CSV_CHARSET,
                    errors="ignore",
                ).title()
                for f in custrows
                if not f[1] and not request.rows[f[0]].hidden
            ]
            if len(scenario_list) > 1:
                r.insert(0, _("scenario"))
            writer.writerow(r)
            fields = [
                request.rows[f[0]]
                for f in custrows
                if not f[1] and not request.rows[f[0]].hidden
            ]
        else:
            # Default settings
            r = [
                force_str(
                    f.title, encoding=settings.CSV_CHARSET, errors="ignore"
                ).title()
                for f in request.rows
                if f.title and not f.hidden and not f.initially_hidden
            ]
            if len(scenario_list) > 1:
                r.insert(0, _("scenario"))
            writer.writerow(r)
            fields = [
                i
                for i in request.rows
                if i.field_name and not i.hidden and not i.initially_hidden
            ]

        # Write a header row
        yield sf.getvalue()

        # Write the report content
        original_database = request.database
        try:
            for scenario in scenario_list:
                request.database = scenario
                if hasattr(request, "query"):
                    delattr(request, "query")

                for row in cls.data_query(request, *args, fields=fields, **kwargs):
                    # Clear the return string buffer
                    sf.seek(0)
                    sf.truncate(0)
                    # Build the return value, encoding all output
                    if hasattr(row, "__getitem__"):
                        r = [
                            cls._getCSVValue(
                                row[f.field_name],
                                field=f,
                                request=request,
                                decimal_separator=decimal_separator,
                                excel_duration_in_days=excel_duration_in_days,
                            )
                            for f in fields
                        ]
                    else:
                        r = [
                            cls._getCSVValue(
                                getattr(row, f.field_name),
                                field=f,
                                request=request,
                                decimal_separator=decimal_separator,
                                excel_duration_in_days=excel_duration_in_days,
                            )
                            for f in fields
                        ]
                    if len(scenario_list) > 1:
                        r.insert(0, scenario_list[scenario])
                    writer.writerow(r)
                    # Return string
                    yield sf.getvalue()
        finally:
            request.database = original_database

    @classmethod
    def getSortName(cls, request):
        """
        Build a jqgrid sort configuration pair sidx and sord:
        For instance:
           ("fieldname1 asc, fieldname2", "desc")
        """
        if request.GET.get("sidx", ""):
            # 1) Sorting order specified on the request
            return (request.GET["sidx"], request.GET.get("sord", "asc"))
        elif request.prefs:
            # 2) Sorting order from the preferences
            sortname = (
                request.prefs.get("sidx", None),
                request.prefs.get("sord", "asc"),
            )
            if sortname[0] and sortname[1]:
                return sortname
        # 3) Default sort order
        if not cls.default_sort:
            return ("", "")
        elif len(cls.default_sort) >= 6:
            return (
                "%s %s, %s %s, %s"
                % (
                    request.rows[cls.default_sort[0]].name,
                    cls.default_sort[1],
                    request.rows[cls.default_sort[2]].name,
                    cls.default_sort[3],
                    request.rows[cls.default_sort[4]].name,
                ),
                cls.default_sort[5],
            )
        elif len(cls.default_sort) >= 4:
            return (
                "%s %s, %s"
                % (
                    request.rows[cls.default_sort[0]].name,
                    cls.default_sort[1],
                    request.rows[cls.default_sort[2]].name,
                ),
                cls.default_sort[3],
            )
        elif len(cls.default_sort) >= 2:
            return (request.rows[cls.default_sort[0]].name, cls.default_sort[1])

    @classmethod
    def _apply_sort(cls, request, query):
        """
        Applies a sort to the query.
        """
        sortname = None
        if request.GET.get("sidx", ""):
            # 1) Sorting order specified on the request
            sortname = "%s %s" % (request.GET["sidx"], request.GET.get("sord", "asc"))
        elif request.prefs:
            # 2) Sorting order from the preferences
            sortname = "%s %s" % (
                request.prefs.get("sidx", ""),
                request.GET.get("sord", "asc"),
            )
        if not sortname or sortname == " asc":
            # 3) Default sort order
            if not cls.default_sort:
                return query
            elif len(cls.default_sort) > 6:
                return query.order_by(
                    (
                        request.rows[cls.default_sort[0]].field_name
                        if cls.default_sort[1] == "asc"
                        else ("-%s" % request.rows[cls.default_sort[0]].field_name)
                    ),
                    (
                        request.rows[cls.default_sort[2]].field_name
                        if cls.default_sort[3] == "asc"
                        else ("-%s" % request.rows[cls.default_sort[2]].field_name)
                    ),
                    (
                        request.rows[cls.default_sort[4]].field_name
                        if cls.default_sort[5] == "asc"
                        else ("-%s" % request.rows[cls.default_sort[4]].field_name)
                    ),
                )
            elif len(cls.default_sort) >= 4:
                return query.order_by(
                    (
                        request.rows[cls.default_sort[0]].field_name
                        if cls.default_sort[1] == "asc"
                        else ("-%s" % request.rows[cls.default_sort[0]].field_name)
                    ),
                    (
                        request.rows[cls.default_sort[2]].field_name
                        if cls.default_sort[3] == "asc"
                        else ("-%s" % request.rows[cls.default_sort[2]].field_name)
                    ),
                )
            elif len(cls.default_sort) >= 2:
                return query.order_by(
                    request.rows[cls.default_sort[0]].field_name
                    if cls.default_sort[1] == "asc"
                    else ("-%s" % request.rows[cls.default_sort[0]].field_name)
                )
            else:
                return query
        else:
            # Validate the field does exist.
            # We only validate the first level field, and not the fields
            # on related models.
            sortargs = []
            for s in sortname.split(","):
                stripped = s.strip()
                if not stripped:
                    continue
                try:
                    sortfield, direction = stripped.split(" ", 1)
                except ValueError:
                    continue
                try:
                    query.order_by(sortfield).query.__str__()
                    if direction.strip() != "desc":
                        sortargs.append(sortfield)
                    else:
                        sortargs.append("-%s" % sortfield)
                except Exception:
                    for r in request.rows:
                        if r.name == sortfield:
                            try:
                                query.order_by(r.field_name).query.__str__()
                                if direction.strip() != "desc":
                                    sortargs.append(r.field_name)
                                else:
                                    sortargs.append("-%s" % r.field_name)
                            except Exception:
                                # Can't sort on this field
                                pass
                            break
            if sortargs:
                return query.order_by(*sortargs)
            else:
                return query

    @classmethod
    def _apply_sort_index(cls, request):
        """
        Build an SQL fragment to sort on: Eg "1 asc, 2 desc"
        """
        sortname = None
        if request.GET.get("sidx", ""):
            # 1) Sorting order specified on the request
            sortname = "%s %s" % (request.GET["sidx"], request.GET.get("sord", "asc"))
        elif request.prefs:
            # 2) Sorting order from the preferences
            sortname = "%s %s" % (
                request.prefs.get("sidx", ""),
                request.prefs.get("sord", "asc"),
            )
        if not sortname or sortname == " asc":
            # 3) Default sort order
            if not cls.default_sort:
                return "1 asc"
            elif len(cls.default_sort) > 6:
                return "%s %s, %s %s, %s %s" % (
                    cls.default_sort[0] + 1,
                    cls.default_sort[1],
                    cls.default_sort[2] + 1,
                    cls.default_sort[3],
                    cls.default_sort[4] + 1,
                    cls.default_sort[5],
                )
            elif len(cls.default_sort) >= 4:
                return "%s %s, %s %s" % (
                    cls.default_sort[0] + 1,
                    cls.default_sort[1],
                    cls.default_sort[2] + 1,
                    cls.default_sort[3],
                )
            elif len(cls.default_sort) >= 2:
                return "%s %s" % (cls.default_sort[0] + 1, cls.default_sort[1])
            else:
                return "1 asc"
        else:
            # Validate the field does exist.
            # We only validate the first level field, and not the fields
            # on related models.
            sortargs = []
            for s in sortname.split(","):
                sortfield, direction = s.strip().split(" ", 1)
                idx = 1
                has_one = False
                for i in request.rows:
                    if i.name == sortfield:
                        sortargs.append(
                            "%s %s" % (idx, "desc" if direction == "desc" else "asc")
                        )
                        if idx == 1:
                            has_one = True
                    idx += 1
            if sortargs:
                if not has_one:
                    sortargs.append("1 asc")
                return ", ".join(sortargs)
            else:
                return "1 asc"

    @classmethod
    def defaultSortString(cls, request):
        if not cls.default_sort:
            return " asc"
        elif len(cls.default_sort) >= 6:
            return "%s %s, %s %s, %s %s" % (
                request.rows[cls.default_sort[0]].name,
                cls.default_sort[1],
                request.rows[cls.default_sort[2]].name,
                cls.default_sort[3],
                request.rows[cls.default_sort[4]].name,
                cls.default_sort[5],
            )
        elif len(cls.default_sort) >= 4:
            return (
                "%s %s, %s %s"
                % (
                    request.rows[cls.default_sort[0]].name,
                    cls.default_sort[1],
                    request.rows[cls.default_sort[2]].name,
                    cls.default_sort[3],
                ),
            )
        elif len(cls.default_sort) >= 2:
            return "%s %s" % (
                request.rows[cls.default_sort[0]].name,
                cls.default_sort[1],
            )
        else:
            return " asc"

    @classmethod
    def get_sort(cls, request):
        try:
            if "sidx" in request.GET:
                # Special case when views have grouping.
                # The group-by column is then added automatically.
                column = request.GET["sidx"]
                comma = column.find(",")
                if comma > 0:
                    column = column[comma + 2 :]
                sort = 1
                ok = False
                for r in request.rows:
                    if r.name == column:
                        ok = True
                        break
                    sort += 1
                if not ok:
                    sort = cls.default_sort[0]
            else:
                sort = cls.default_sort[0]
        except Exception:
            sort = cls.default_sort[0]
        if request.GET.get("sord", None) == "desc" or cls.default_sort[1] == "desc":
            return "%s desc" % sort
        else:
            return "%s asc" % sort

    @classmethod
    def data_query(cls, request, *args, fields=None, page=None, **kwargs):
        if not fields:
            raise Exception("No fields provided as argument")
        if not hasattr(request, "query"):
            if callable(cls.basequeryset):
                request.query = cls.filter_items(
                    request, cls.basequeryset(request, *args, **kwargs), False
                ).using(request.database)
            else:
                request.query = cls.filter_items(request, cls.basequeryset).using(
                    request.database
                )
        query = cls._apply_sort(request, request.query)
        if page:
            # Display a single page
            cnt = (page - 1) * request.pagesize + 1
            if hasattr(cls, "query"):
                return cls.query(request, query[cnt - 1 : cnt + request.pagesize])
            else:
                return query[cnt - 1 : cnt + request.pagesize].values(*fields)
        else:
            limit = getattr(request, "limit", 0)
            if limit:
                # Kanban views have no pages, but still limit the number of cards
                if hasattr(cls, "query"):
                    return cls.query(request, query[:limit])
                else:
                    return query[:limit].values(*fields)
            else:
                # No size limit on the query results
                if hasattr(cls, "query"):
                    return cls.query(request, query)
                else:
                    fields = [i.field_name for i in request.rows if i.field_name]
                    return query.values(*fields)

    @classmethod
    def count_query(cls, request, *args, **kwargs):
        if not hasattr(request, "query"):
            if callable(cls.basequeryset):
                request.query = cls.filter_items(
                    request, cls.basequeryset(request, *args, **kwargs), False
                ).using(request.database)
            else:
                request.query = cls.filter_items(request, cls.basequeryset).using(
                    request.database
                )

        tmp = request.query.query.get_compiler(request.database).as_sql(
            with_col_aliases=False
        )
        with connections[request.database].cursor() as cursor:
            cursor.execute("select count(*) from (" + tmp[0] + ") t_subquery", tmp[1])
            cache_val = cursor.fetchone()[0]
            return cache_val

    @classmethod
    def _generate_json_data(cls, request, *args, **kwargs):
        request.prefs = request.user.getPreference(
            cls.getKey(request, *args, **kwargs), database=request.database
        )
        recs = cls.count_query(request, *args, **kwargs)
        if "rows" in request.GET:
            request.pagesize = int(request.GET["rows"])
        total_pages = math.ceil(float(recs) / request.pagesize)
        page = request.GET.get("page", 1)
        if page is not None:
            page = int(page)
            if page > total_pages:
                page = total_pages
            if page < 1:
                page = 1

        yield '{"total":%d,\n' % total_pages
        if page:
            yield '"page":%d,\n' % page
        yield '"records":%d,\n' % recs
        if hasattr(cls, "extraJSON"):
            # Hook to insert extra fields to the json
            tmp = cls.extraJSON(request)
            if tmp:
                yield tmp
        yield '"rows":[\n'

        # GridReport
        first = True
        fields = [i.field_name for i in request.rows if i.field_name]
        for i in cls.data_query(request, *args, fields=fields, page=page, **kwargs):
            if first:
                r = ["{"]
                first = False
            else:
                r = [",\n{"]
            first2 = True
            for f in request.rows:
                if not f.name:
                    continue
                s = cls._getJSONValue(i[f.field_name], field=f, request=request)
                if first2:
                    r.append('"%s":%s' % (f.name, s))
                    first2 = False
                elif i[f.field_name] is not None:
                    r.append(', "%s":%s' % (f.name, s))
            r.append("}")
            yield "".join(r)
        yield "\n]}\n"

    @classmethod
    def post(cls, request, *args, **kwargs):
        if len(request.FILES) > 0:

            # confirm there is enough storage to proceed
            maxstorage = getattr(settings, "MAXSTORAGE", 0) or 0
            if maxstorage:
                storageUsage = round(getStorageUsage() / 1024 / 1024)
                if storageUsage > maxstorage:
                    return HttpResponseForbidden(
                        """
                        Storage quota exceeded: %sMB used out of %sMB available.<br>
                        Please <a class="text-decoration-underline" href="%s/docs/current/doc/installation-guide/setting-disk-space-quotas.html" target="_blank">free some disk space</a>
                        and try again.
                        """
                        % (storageUsage, maxstorage, settings.DOCUMENTATION_URL)
                    )

            # Note: the detection of the type of uploaded file depends on the
            # browser setting the right mime type of the file.
            csvcount = 0
            xlscount = 0
            for filename, file in request.FILES.items():
                if file.content_type in (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel.sheet.macroenabled.12",
                ):
                    xlscount += 1
                elif filename.endswith(".xls"):
                    return HttpResponseNotFound(
                        _(
                            "Files in the old .XLS excel format can't be read.<br>"
                            "Please convert them to the new .XLSX format."
                        )
                    )
                else:
                    csvcount += 1

            if csvcount == 0:
                # Uploading a spreadsheet file
                return StreamingHttpResponse(
                    content_type="text/plain; charset=%s" % settings.DEFAULT_CHARSET,
                    streaming_content=cls.parseSpreadsheetUpload(request),
                )
            elif xlscount == 0:
                # Uploading a CSV file
                return StreamingHttpResponse(
                    content_type="text/plain; charset=%s" % settings.DEFAULT_CHARSET,
                    streaming_content=cls.parseCSVupload(request),
                )
            else:  # mixed files
                return HttpResponseNotFound("Files must have the same type.")
        else:
            # Saving after inline edits
            return cls.parseJSONupload(request)

    @classmethod
    def _validate_crosses(cls, request, prefs):
        cross_idx = []
        for i in prefs:
            try:
                num = int(i)
                if num < len(request.crosses) and request.crosses[num][1].get(
                    "visible", True
                ):
                    cross_idx.append(num)
            except ValueError:
                for j in range(len(request.crosses)):
                    if request.crosses[j][0] == i and request.crosses[j][1].get(
                        "visible", True
                    ):
                        cross_idx.append(j)
        return cross_idx

    @classmethod
    def getScenarios(cls, request, *args, **kwargs):
        scenario_permissions = []
        if len(request.user.scenarios) > 1:
            original_database = request.database
            for scenario in request.user.scenarios:
                try:
                    # request.database needs to be changed for has_perm to work properly
                    request.database = scenario.name
                    if hasattr(request, "query"):
                        delattr(request, "query")

                    user = User.objects.using(scenario.name).get(
                        username=request.user.username
                    )

                    if cls.has_permission(user):
                        scenario_permissions.append(
                            [
                                scenario.name,
                                (
                                    scenario.description
                                    if scenario.description
                                    else scenario.name
                                ),
                                1 if request.database == original_database else 0,
                            ]
                        )
                except Exception:
                    pass

            # reverting to original request database as permissions are checked
            request.database = original_database
        return scenario_permissions

    @classmethod
    def get(cls, request, *args, **kwargs):
        # Pick up the list of time buckets
        if cls.hasTimeBuckets:
            cls.getBuckets(request, args, kwargs)
            bucketnames = Bucket.objects.using(request.database)
            if cls.maxBucketLevel:
                if callable(cls.maxBucketLevel):
                    maxlvl = cls.maxBucketLevel(request)
                    bucketnames = bucketnames.filter(level__lte=maxlvl)
                else:
                    bucketnames = bucketnames.filter(level__lte=cls.maxBucketLevel)
            if cls.minBucketLevel:
                if callable(cls.minBucketLevel):
                    minlvl = cls.minBucketLevel(request)
                    bucketnames = bucketnames.filter(level__gte=minlvl)
                else:
                    bucketnames = bucketnames.filter(level__gte=cls.minBucketLevel)
            bucketnames = bucketnames.order_by("-level").values_list("name", flat=True)
        else:
            bucketnames = None
        fmt = request.GET.get("format", None)
        reportkey = cls.getKey(request, *args, **kwargs)
        request.prefs = request.user.getPreference(reportkey, database=request.database)
        if request.prefs:
            kwargs["preferences"] = request.prefs

        # check if duplicate icon should be hidden
        if hasattr(cls, "model") and cls.model:
            if (
                hasattr(cls.model._meta, "unique_together")
                and len(cls.model._meta.unique_together) > 0
            ):
                cls.canDuplicate = False

        if not fmt:
            # Return HTML page
            if not hasattr(request, "crosses"):
                cross_idx = None
                cross_list = None
            elif request.prefs and "crosses" in request.prefs:
                cross_idx = str(
                    cls._validate_crosses(request, request.prefs["crosses"])
                )
                cross_list = cls._render_cross(request)
            else:
                cross_idx = str(
                    [
                        i
                        for i in range(len(request.crosses))
                        if request.crosses[i][1].get("visible", True)
                        and not request.crosses[i][1].get("initially_hidden", False)
                    ]
                )
                cross_list = cls._render_cross(request)
            if args and not cls.new_arg_logic:
                mode = "table"
            else:
                mode = request.GET.get("mode", None)
                if mode:
                    # Store the mode passed in the URL on the session to remember for the next report
                    request.session["mode"] = mode
                else:
                    # Pick up the mode from the session
                    mode = request.session.get("mode", "graph")
            is_popup = "_popup" in request.GET
            sidx, sord = cls.getSortName(request)

            autofilter = "noautofilter" not in request.GET and cls.autofilter
            filters = cls.getQueryString(request)
            if not filters and request.prefs and autofilter:
                # Inherit the filter settings from the preferences
                filters = request.prefs.get("filter", None)
            if request.prefs and autofilter:
                page = request.prefs.get("page", 1) or 1
            else:
                page = 1
            context = {
                "reportclass": cls,
                "title": (
                    _("%(title)s for %(entity)s")
                    % {"title": force_str(cls.title), "entity": force_str(args[0])}
                    if args and args[0]
                    else cls.title
                ),
                "post_title": cls.post_title,
                "preferences": request.prefs,
                "reportkey": reportkey,
                "colmodel": cls._render_colmodel(
                    request, is_popup, request.prefs, mode, *args, **kwargs
                ),
                "cross_idx": cross_idx,
                "cross_list": cross_list,
                "object_id": args and quote(args[0]) or None,
                "page": page,
                "sord": sord,
                "sidx": sidx,
                "default_sort": cls.defaultSortString(request),
                "is_popup": is_popup,
                "filters": json.loads(filters) if filters else None,
                "args": args,
                "bucketnames": bucketnames,
                "model": cls.model,
                # Note: we don't check here whether the user has the required priviliges
                # in the requested scenario. This check is only done when the export is
                # executed.
                "scenario_permissions": [
                    [
                        i.name,
                        i.tag,
                        1 if i.name == request.database else 0,
                    ]
                    for i in request.user.scenarios
                ],
                "hasaddperm": cls.editable
                and cls.model
                and (
                    request.user.has_perm(
                        "%s.%s"
                        % (
                            cls.model._meta.app_label,
                            get_permission_codename("add", cls.model._meta),
                        )
                    )
                    or "add" not in cls.model._meta.default_permissions
                ),
                "hasdeleteperm": cls.editable
                and cls.model
                and (
                    request.user.has_perm(
                        "%s.%s"
                        % (
                            cls.model._meta.app_label,
                            get_permission_codename("delete", cls.model._meta),
                        )
                    )
                    or "delete" not in cls.model._meta.default_permissions
                ),
                "haschangeperm": cls.editable
                and cls.model
                and (
                    request.user.has_perm(
                        "%s.%s"
                        % (
                            cls.model._meta.app_label,
                            get_permission_codename("change", cls.model._meta),
                        )
                    )
                    or "change" not in cls.model._meta.default_permissions
                ),
                "active_tab": "plan",
                "mode": mode,
                "actions": cls.actions,
            }
            for k, v in cls.extra_context(request, *args, **kwargs).items():
                context[k] = v
            return render(request, cls.template, context)
        elif fmt == "json":
            # Return JSON data to fill the grid.
            response = StreamingHttpResponse(
                content_type="application/json; charset=%s" % settings.DEFAULT_CHARSET,
                streaming_content=cls._generate_json_data(request, *args, **kwargs),
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt in ("spreadsheetlist", "spreadsheettable", "spreadsheet"):
            scenarios = request.GET.get("scenarios", None)
            scenario_list = scenarios.split(",") if scenarios else [request.database]
            # Make sure the user has the requiredd permissions in the requested scenarios!
            if scenarios:
                scenario_permissions = cls.getScenarios(request, *args, **kwargs)
                scenario_list = {
                    t[0]: t[1] for t in scenario_permissions if t[0] in scenario_list
                }

            # Return an excel spreadsheet
            output = BytesIO()
            cls._generate_spreadsheet_data(
                request, scenario_list, output, *args, **kwargs
            )
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content=output.getvalue(),
            )
            # Filename parameter is encoded as specified in rfc5987
            if callable(cls.title):
                title = cls.title(request, *args, **kwargs)
            elif hasattr(cls, "title"):
                title = cls.title
            else:
                title = cls.model._meta.verbose_name_plural if cls.model else cls.title
            response["Content-Disposition"] = (
                "attachment; filename*=utf-8''%s.xlsx"
                % urllib.parse.quote(force_str(title))
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt in ("csvlist", "csvtable", "csv"):
            scenarios = request.GET.get("scenarios", None)
            scenario_list = scenarios.split(",") if scenarios else [request.database]
            # Make sure scenarios are in the scenario_permissions list
            if scenarios:
                scenario_list = {
                    t[0]: t[1] for t in scenario_permissions if t[0] in scenario_list
                }

            # Return CSV data to export the data
            response = StreamingHttpResponse(
                content_type="text/csv; charset=%s" % settings.CSV_CHARSET,
                streaming_content=cls._generate_csv_data(
                    request, scenario_list, *args, **kwargs
                ),
            )
            # Filename parameter is encoded as specified in rfc5987
            if callable(cls.title):
                title = cls.title(request, *args, **kwargs)
            elif hasattr(cls, "title"):
                title = cls.title
            else:
                title = cls.model._meta.verbose_name_plural if cls.model else cls.title
            response["Content-Disposition"] = (
                "attachment; filename*=utf-8''%s.csv"
                % urllib.parse.quote(force_str(title))
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt == "kanban":
            response = StreamingHttpResponse(
                content_type="application/json; charset=%s" % settings.DEFAULT_CHARSET,
                streaming_content=cls._generate_kanban_data(request, *args, **kwargs),
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt == "gantt":
            response = StreamingHttpResponse(
                content_type="application/json; charset=%s" % settings.DEFAULT_CHARSET,
                streaming_content=cls._generate_gantt_data(request, *args, **kwargs),
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        elif fmt == "calendar":
            response = StreamingHttpResponse(
                content_type="application/json; charset=%s" % settings.DEFAULT_CHARSET,
                streaming_content=cls._generate_calendar_data(request, *args, **kwargs),
            )
            response["Cache-Control"] = "no-cache, no-store"
            return response
        else:
            raise Http404("Unknown format type")

    @classmethod
    def _generate_kanban_data(cls, request, *args, **kwargs):
        raise Http404("This report doesn't support the kanban format")

    @classmethod
    def _generate_gantt_data(cls, request, *args, **kwargs):
        raise Http404("This report doesn't support the gantt format")

    @classmethod
    def _generate_calendar_data(cls, request, *args, **kwargs):
        raise Http404("This report doesn't support the calendar format")

    @classmethod
    def parseJSONupload(cls, request):
        # Check permissions
        if not cls.model or not cls.editable:
            return HttpResponseForbidden(_("Permission denied"))
        permname = get_permission_codename("change", cls.model._meta)
        if not request.user.has_perm("%s.%s" % (cls.model._meta.app_label, permname)):
            return HttpResponseForbidden(_("Permission denied"))

        # Loop over the data records
        resp = HttpResponse()
        ok = True
        data = json.JSONDecoder().decode(
            request.read().decode(request.encoding or settings.DEFAULT_CHARSET)
        )
        with transaction.atomic(using=request.database, savepoint=False):
            content_type_id = ContentType.objects.get_for_model(
                cls.model, for_concrete_model=False
            ).pk
            for rec in data:
                if "update" in rec:
                    # Bulk update
                    fields = {}
                    for f, v in data["update"].get("fields", {}).items():
                        if v is not None:
                            v = v.strip()
                        r = cls._getRowByName(request, f)
                        if not r.editable:
                            continue
                        if isinstance(r, (GridFieldCurrency, GridFieldNumber)):
                            try:
                                fields[f] = float(v) if v is not None else None
                            except Exception:
                                ok = False
                                resp.write("Invalid number %s<br>" % v)
                        elif isinstance(r, GridFieldInteger):
                            try:
                                fields[f] = int(v) if v is not None else None
                            except Exception:
                                ok = False
                                resp.write("Invalid integer %s<br>" % v)
                        elif isinstance(r, GridFieldChoice):
                            if v is None:
                                fields[f] = None
                            elif False and v.lower() in [
                                force_str(i[1]).lower() for i in r.choices
                            ]:
                                fields[f] = [
                                    x
                                    for (x, y) in r.choices
                                    if force_str(y).lower() == v.lower()
                                ][0]
                            else:
                                ok = False
                                resp.write(
                                    "Invalid choice %s: expected %s<br>"
                                    % (v, ", ".join([c[0] for c in r.choices]))
                                )
                        elif isinstance(r, GridFieldBool):
                            if v is not None:
                                fields[f] = v.lower() in [
                                    "0",
                                    "false",
                                    force_str(_("false")),
                                ]
                            else:
                                ok = False
                                resp.write("Bool field can't be empty<br>")
                        elif isinstance(r, GridFieldBoolNullable):
                            if v is not None:
                                fields[f] = v.lower() in [
                                    "0",
                                    "false",
                                    force_str(_("false")),
                                ]
                            else:
                                fields[f] = None
                        elif r.formatter == "detail":
                            try:
                                modelname = json.loads("{%s}" % r)["role"].split("/")
                                model = apps.get_model(modelname[0], modelname[1])
                                fields[f] = (
                                    model.objects.using(request.database).get(pk=v)
                                    if v
                                    else None
                                )
                            except Exception:
                                ok = False
                                resp.write(
                                    "New value doesn't exist in related table<br>"
                                )
                        elif isinstance(r, (GridFieldDateTime, GridFieldLocalDateTime)):
                            try:
                                fields[f] = (
                                    parseLocalizedDateTime(v) if v is not None else None
                                )
                            except Exception:
                                ok = False
                                resp.write("Invalid datetime format<br>")
                        elif isinstance(r, GridFieldDate):
                            try:
                                fields[f] = (
                                    parseLocalizedDate(v) if v is not None else None
                                )
                            except Exception:
                                ok = False
                                resp.write("Invalid date format<br>")
                        elif isinstance(r, GridFieldTime):
                            if v is None:
                                fields[f] = v
                            else:
                                try:
                                    fields[f] = parse_time(v)
                                    if fields[f] is None:
                                        raise Exception
                                except Exception:
                                    ok = False
                                    resp.write("Invalid time format<br>")
                        elif isinstance(r, GridFieldDuration):
                            if v is None:
                                fields[f] = v
                            else:
                                try:
                                    fields[f] = parse_duration(v)
                                    if fields[f] is None:
                                        raise Exception
                                except Exception:
                                    ok = False
                                    resp.write("Invalid duration format<br>")
                        else:
                            fields[f] = v
                    if fields and ok:
                        sid = transaction.savepoint(using=request.database)
                        try:
                            comment = "Changed %s." % get_text_list(
                                list(fields.keys()), "and"
                            )
                            if "pk" in data["update"]:
                                cls.model.objects.all().using(request.database).filter(
                                    pk__in=data["update"]["pk"]
                                ).update(**fields)
                                for o in data["update"]["pk"]:
                                    Comment(
                                        user_id=request.user.id,
                                        content_type_id=content_type_id,
                                        object_pk=force_str(o),
                                        object_repr=o[:200],
                                        type="change",
                                        comment=comment,
                                    ).save(using=request.database)
                            else:
                                objs = cls.model.objects.all().using(request.database)
                                flt = cls._get_q_filter(
                                    request, data["update"].get("filter", [])
                                )
                                if flt:
                                    objs = objs.filter(flt)
                                objs.update(**fields)
                                for k in objs:
                                    Comment(
                                        user_id=request.user.id,
                                        content_type_id=content_type_id,
                                        object_pk=force_str(k.pk),
                                        object_repr=force_str(k)[:200],
                                        type="change",
                                        comment=comment,
                                    ).save(using=request.database)
                            transaction.savepoint_commit(sid)
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            ok = False
                            resp.write(escape(e))
                            resp.write("<br>")
                elif "delete" in rec:
                    # Deleting records
                    for key in rec["delete"]:
                        sid = transaction.savepoint(using=request.database)
                        try:
                            obj = cls.model.objects.using(request.database).get(pk=key)
                            Comment(
                                user_id=request.user.id,
                                content_type_id=content_type_id,
                                object_pk=force_str(key),
                                object_repr=force_str(obj)[:200],
                                type="delete",
                                comment="Deleted %s." % force_str(obj),
                            ).save(using=request.database)
                            obj.delete()
                            transaction.savepoint_commit(sid)
                        except cls.model.DoesNotExist:
                            transaction.savepoint_rollback(sid)
                            ok = False
                            resp.write(escape(_("Can't find %s" % key)))
                            resp.write("<br>")
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            ok = False
                            resp.write(escape(e))
                            resp.write("<br>")
                elif "copy" in rec:
                    # Copying records
                    for key in rec["copy"]:
                        sid = transaction.savepoint(using=request.database)
                        try:
                            obj = cls.model.objects.using(request.database).get(pk=key)
                            orig_repr = force_str(obj)
                            if isinstance(cls.model._meta.pk, CharField):
                                # The primary key is a string
                                copy_index = 1
                                basekey = key.partition(" - copy #")[0]
                                while True:
                                    obj.pk = "%s - copy #%s" % (
                                        basekey,
                                        copy_index,
                                    )
                                    if (
                                        cls.model.objects.using(request.database)
                                        .filter(pk=obj.pk)
                                        .exists()
                                    ):
                                        copy_index += 1
                                    else:
                                        break
                            elif isinstance(cls.model._meta.pk, AutoField):
                                # The primary key is an auto-generated number
                                obj.pk = None
                            else:
                                raise Exception(
                                    _("Can't copy %s") % cls.model._meta.app_label
                                )
                            obj.save(using=request.database, force_insert=True)
                            Comment(
                                user_id=request.user.pk,
                                content_type_id=content_type_id,
                                object_pk=obj.pk,
                                object_repr=force_str(obj)[:200],
                                type="add",
                                comment="Copied from %s." % orig_repr,
                            ).save(using=request.database)
                            transaction.savepoint_commit(sid)
                        except cls.model.DoesNotExist:
                            transaction.savepoint_rollback(sid)
                            ok = False
                            resp.write(escape(_("Can't find %s" % key)))
                            resp.write("<br>")
                        except Exception as e:
                            transaction.savepoint_rollback(sid)
                            ok = False
                            resp.write(escape(e))
                            resp.write("<br>")
                else:
                    # Editing records
                    pk = rec["id"]
                    sid = transaction.savepoint(using=request.database)
                    try:
                        obj = cls.model.objects.using(request.database).get(
                            pk=rec["id"]
                        )
                        del rec["id"]
                        for i in rec:
                            if (
                                rec[i] == "\xa0"
                            ):  # Workaround for Jqgrid issue: date field can't be set to blank
                                rec[i] = None
                        # Assure all unique-together fields are included
                        flds = set(rec.keys())
                        for c in cls.model._meta.unique_together:
                            for cf in c:
                                if cf in flds:
                                    for cf2 in c:
                                        if cf2 not in flds:
                                            rec[cf2] = str(getattr(obj, cf2))
                                            flds.add(cf2)
                                    break
                        if hasattr(cls.model, "getModelForm"):
                            UploadForm = cls.model.getModelForm(
                                tuple(flds), database=request.database
                            )
                        else:
                            UploadForm = modelform_factory(
                                cls.model,
                                fields=tuple(flds),
                                formfield_callback=lambda f: (
                                    isinstance(f, RelatedField)
                                    and f.formfield(using=request.database)
                                )
                                or (
                                    isinstance(f, DateTimeField)
                                    and f.formfield(
                                        input_formats=settings.DATETIME_INPUT_FORMATS
                                    )
                                )
                                or (
                                    isinstance(f, DateField)
                                    and f.formfield(
                                        input_formats=settings.DATE_INPUT_FORMATS
                                    )
                                )
                                or f.formfield(),
                            )
                        form = UploadForm(rec, instance=obj)
                        if not form.is_valid():
                            raise ValueError
                        elif form.has_changed():
                            obj = form.save(commit=False)
                            if not hasattr(obj, "skipsave"):
                                obj.save(using=request.database)
                            Comment(
                                user_id=request.user.pk,
                                content_type_id=content_type_id,
                                object_pk=obj.pk,
                                object_repr=force_str(obj)[:200],
                                type="change",
                                comment="Changed %s."
                                % get_text_list(form.changed_data, "and"),
                            ).save(using=request.database)
                        transaction.savepoint_commit(sid)
                    except cls.model.DoesNotExist:
                        transaction.savepoint_rollback(sid)
                        ok = False
                        resp.write(escape(_("Can't find %s" % pk)))
                        resp.write("<br>")
                    except (ValidationError, ValueError):
                        transaction.savepoint_rollback(sid)
                        ok = False
                        for error in form.non_field_errors():
                            resp.write(escape("%s: %s" % (pk, error)))
                            resp.write("<br>")
                        for field in form:
                            for error in field.errors:
                                resp.write(
                                    escape(
                                        "%s %s: %s: %s"
                                        % (
                                            obj.pk,
                                            field.name,
                                            rec.get(
                                                field.name,
                                                getattr(obj, field.name),
                                            ),
                                            error,
                                        )
                                    )
                                )
                                resp.write("<br>")
                    except Exception as e:
                        transaction.savepoint_rollback(sid)
                        ok = False
                        resp.write(escape(e))
                        resp.write("<br>")

        # Update the hierachy
        if issubclass(cls.model, HierarchyModel):
            cls.model.rebuildHierarchy(database=request.database)

        if ok:
            resp.write("OK")
        resp.status_code = ok and 200 or 500
        return resp

    @staticmethod
    def dependent_models(m, found):
        """An auxilary method that constructs a set of all dependent models"""
        for f in m._meta.get_fields():
            if (
                f.is_relation
                and f.auto_created
                and f.related_model != m
                and f.related_model not in found
            ):
                for sub in f.related_model.__subclasses__():
                    # if sub not in found:
                    found.update([sub])
                found.update([f.related_model])
                GridReport.dependent_models(f.related_model, found)

    @staticmethod
    def sort_models(models):
        # Inject additional dependencies that are not reflected in database constraints
        parameter_model = None
        for m in models:
            if m[1] == Parameter:
                parameter_model = m
            for e in getattr(m[1], "extra_dependencies", []):
                for m2 in models:
                    if m2[1] == e:
                        m2[3].update([m[1]])

        # Sort the list of models, based on dependencies between models
        models.sort(key=lambda m: (m[1].__name__, m[0].upper()))
        cnt = len(models)
        ok = False
        while not ok:
            ok = True
            for i in range(cnt):
                j = i + 1
                while j < cnt and ok:
                    if models[i][1] != models[j][1] and models[i][1] in models[j][3]:
                        i_base = models[i][1].__base__
                        if i_base == Model or i_base._meta.abstract:
                            i_base = None
                        j_base = models[j][1].__base__
                        if j_base == Model or j_base._meta.abstract:
                            j_base = None
                        if i_base == j_base and i_base and j_base:
                            j += 1
                            continue
                        if i_base == models[j][1] or j_base == models[i][1]:
                            j += 1
                            continue
                        models.append(models.pop(i))
                        ok = False
                        break
                    elif (
                        models[i][1] == models[j][1]
                        and models[i][0].upper() > models[j][0].upper()
                    ):
                        models[i], models[j] = models[j], models[i]
                        ok = False
                    j += 1
        # Assure parameters are read first
        if parameter_model:
            models.insert(0, models.pop(models.index(parameter_model)))
        return models

    @classmethod
    def erase(cls, request):
        # Build a list of dependencies
        deps = set([cls.model])
        # Special case for MO/PO/DO/DLVR that cannot be truncated
        if cls.model.__name__ not in (
            "PurchaseOrder",
            "ManufacturingOrder",
            "DistributionOrder",
            "DeliveryOrder",
        ):
            GridReport.dependent_models(cls.model, deps)

        # Check the delete permissions for all related objects
        for m in deps:
            permname = get_permission_codename("delete", m._meta)
            if not request.user.has_perm("%s.%s" % (m._meta.app_label, permname)):
                return format_lazy(
                    "{}:{}", m._meta.verbose_name, _("Permission denied")
                )

        # Delete the data records
        cursor = connections[request.database].cursor()
        with transaction.atomic(using=request.database):
            sql_list = []
            containsOperationPlan = any(m.__name__ == "OperationPlan" for m in deps)
            for m in deps:
                if "getDeleteStatements" in dir(m) and not containsOperationPlan:
                    sql_list.extend(m.getDeleteStatements())
                else:
                    sql_list = connections[request.database].ops.sql_flush(
                        no_style(), [m._meta.db_table for m in deps]
                    )
            for sql in sql_list:
                cursor.execute(sql)
            # Erase comments and history
            content_ids = [
                ContentType.objects.get_for_model(m, for_concrete_model=False)
                for m in deps
            ]
            Comment.objects.filter(content_type__in=content_ids).delete()
            # Prepare message
            for m in deps:
                messages.add_message(
                    request,
                    messages.INFO,
                    _("Erasing data from %(model)s")
                    % {"model": force_str(m._meta.verbose_name)},
                )

        # Finished successfully
        return None

    @classmethod
    def parseCSVupload(cls, request):
        """
        This method reads CSV data from a string (in memory) and creates or updates
        the database records.
        The data must follow the following format:
        - the first row contains a header, listing all field names
        - a first character # marks a comment line
        - empty rows are skipped
        """
        # Check permissions
        if not cls.model:
            yield "<div>%s</div>" % _("Invalid upload request")
            return
        permname = get_permission_codename("add", cls.model._meta)
        if not cls.editable or not request.user.has_perm(
            "%s.%s" % (cls.model._meta.app_label, permname)
        ):
            yield "<div>%s</div>" % _("Permission denied")
            return

        # Choose the right delimiter and language
        delimiter = (
            get_format("DECIMAL_SEPARATOR", request.LANGUAGE_CODE, True) == ","
            and ";"
            or ","
        )
        if translation.get_language() != request.LANGUAGE_CODE:
            translation.activate(request.LANGUAGE_CODE)

        # Handle the complete upload as a single database transaction
        try:
            with transaction.atomic(using=request.database):
                # Erase all records and related tables
                if "erase" in request.POST:
                    returnvalue = cls.erase(request)
                    if returnvalue:
                        yield format_lazy("<div>{}</div>", returnvalue)
                        return

                yield (
                    '<div class="table-responsive">'
                    '<table class="table table-sm" style="white-space: nowrap"><tbody>'
                )

                for filename, file in request.FILES.items():
                    numerrors = 0
                    numwarnings = 0
                    firsterror = True
                    yield '<tr><th colspan="5">%s<div class="recordcount float-right"></div></th></tr>' % filename
                    data = EncodedCSVReader(file, delimiter=delimiter)
                    for error in parseCSVdata(
                        cls.model,
                        data,
                        user=request.user,
                        database=request.database,
                        ping=True,
                    ):
                        if error[0] == logging.DEBUG:
                            # Yield some result so we can detect disconnect clients and interrupt the upload
                            yield "<tr class='hidden' data-cnt='%s'>" % error[1]
                            continue
                        if firsterror and error[0] in (logging.ERROR, logging.WARNING):
                            yield '<tr><th class="sr-only">%s</th><th>%s</th><th>%s</th><th>%s</th><th>%s%s%s</th></tr>' % (
                                capfirst(_("worksheet")),
                                capfirst(_("row")),
                                capfirst(_("field")),
                                capfirst(_("value")),
                                capfirst(_("error")),
                                " / ",
                                capfirst(_("warning")),
                            )
                            firsterror = False
                        if error[0] == logging.ERROR:
                            yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                cls.model._meta.verbose_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                capfirst(_("error")),
                                error[4],
                            )
                            numerrors += 1
                        elif error[1] == logging.WARNING:
                            yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                cls.model._meta.verbose_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                capfirst(_("warning")),
                                error[4],
                            )
                            numwarnings += 1
                        else:
                            yield '<tr class=%s><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                                "danger" if numerrors > 0 else "success",
                                cls.model._meta.verbose_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                error[4],
                            )
                yield "</tbody></table></div>"

            # Update the hierachy
            if issubclass(cls.model, HierarchyModel):
                cls.model.rebuildHierarchy(database=request.database)

            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(
                database=request.database,
                url=(
                    "%s://%s"
                    % ("https" if request.is_secure() else "http", request.get_host())
                    if request
                    else None
                ),
            )
        except GeneratorExit:
            logging.warning("Connection Aborted")
        except NameError:
            pass

    @classmethod
    def parseSpreadsheetUpload(cls, request):
        """
        This method reads a spreadsheet file (in memory) and creates or updates
        the database records.
        The data must follow the following format:
          - only the first tab in the spreadsheet is read
          - the first row contains a header, listing all field names
          - a first character # marks a comment line
          - empty rows are skipped
        """
        # Check permissions
        if not cls.model:
            yield "<div>%s</div>" % _("Invalid upload request")
            return
        permname = get_permission_codename("add", cls.model._meta)
        if not cls.editable or not request.user.has_perm(
            "%s.%s" % (cls.model._meta.app_label, permname)
        ):
            yield "<div>%s</div>" % _("Permission denied")
            return

        # Choose the right language
        if translation.get_language() != request.LANGUAGE_CODE:
            translation.activate(request.LANGUAGE_CODE)

        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue(
                "excel_duration_in_days", request.database, "false"
            ).lower()
            == "true"
        )

        # Handle the complete upload as a single database transaction
        try:
            with transaction.atomic(using=request.database):
                # Erase all records and related tables
                if "erase" in request.POST:
                    returnvalue = cls.erase(request)
                    if returnvalue:
                        yield '<br><samp style="padding-left: 15px;">%s</samp><br>' % returnvalue
                        raise StopIteration

                # Header in output
                yield (
                    '<div class="table-responsive">'
                    '<table class="table table-sm" style="white-space: nowrap"><tbody>'
                )

                for filename, file in request.FILES.items():
                    numerrors = 0
                    numwarnings = 0
                    firsterror = True
                    yield '<tr"><th colspan="5">%s %s<div class="recordcount float-right"></div></th></tr>' % (
                        capfirst(_("file")),
                        filename,
                    )

                    # Loop through the data records
                    wb = load_workbook(filename=file, data_only=True)
                    numsheets = len(wb.sheetnames)

                    for ws_name in wb.sheetnames:
                        rowprefix = "" if numsheets == 1 else "%s " % ws_name
                        ws = wb[ws_name]
                        if not ws.sheet_state == "visible":
                            # Skip hidden sheets
                            continue
                        for error in parseExcelWorksheet(
                            cls.model,
                            ws,
                            user=request.user,
                            database=request.database,
                            ping=True,
                            excel_duration_in_days=excel_duration_in_days,
                        ):
                            if error[0] == logging.DEBUG:
                                # Yield some result so we can detect disconnect clients and interrupt the upload
                                yield "<tr class='hidden' data-cnt='%s'>" % error[1]
                                continue
                            if firsterror and error[0] in (
                                logging.ERROR,
                                logging.WARNING,
                            ):
                                yield '<tr><th class="sr-only">%s</th><th>%s</th><th>%s</th><th>%s</th><th>%s%s%s</th></tr>' % (
                                    capfirst(_("worksheet")),
                                    capfirst(_("row")),
                                    capfirst(_("field")),
                                    capfirst(_("value")),
                                    capfirst(_("error")),
                                    " / ",
                                    capfirst(_("warning")),
                                )
                                firsterror = False
                            if error[0] == logging.ERROR:
                                yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                    cls.model._meta.verbose_name,
                                    error[1] if error[1] else "",
                                    (
                                        "%s%s" % (rowprefix, error[2] or "")
                                        if rowprefix or error[2]
                                        else ""
                                    ),
                                    error[3] if error[3] else "",
                                    capfirst(_("error")),
                                    error[4],
                                )
                                numerrors += 1
                            elif error[1] == logging.WARNING:
                                yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                    cls.model._meta.verbose_name,
                                    error[1] if error[1] else "",
                                    (
                                        "%s%s" % (rowprefix, error[2] or "")
                                        if rowprefix or error[2]
                                        else ""
                                    ),
                                    error[3] if error[3] else "",
                                    capfirst(_("warning")),
                                    error[4],
                                )
                                numwarnings += 1
                            else:
                                yield '<tr class=%s><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                                    "danger" if numerrors > 0 else "success",
                                    cls.model._meta.verbose_name,
                                    error[1] if error[1] else "",
                                    (
                                        "%s%s" % (rowprefix, error[2] or "")
                                        if rowprefix or error[2]
                                        else ""
                                    ),
                                    error[3] if error[3] else "",
                                    error[4],
                                )
                yield "</tbody></table></div>"

            # Update the hierachy
            if issubclass(cls.model, HierarchyModel):
                cls.model.rebuildHierarchy(database=request.database)

            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(
                database=request.database,
                url=(
                    "%s://%s"
                    % ("https" if request.is_secure() else "http", request.get_host())
                    if request
                    else None
                ),
            )
        except GeneratorExit:
            logger.warning("Connection Aborted")
        except NameError:
            pass

    @classmethod
    def _getRowByName(cls, request, name):
        if not hasattr(cls, "_rowsByName"):
            cls._rowsByName = {}
            for i in request.rows:
                cls._rowsByName[i.name] = i
                if i.field_name != i.name:
                    cls._rowsByName[i.field_name] = i
        return cls._rowsByName[name]

    @staticmethod
    def _filter_ne(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(
            reportrow, (GridFieldCurrency, GridFieldInteger, GridFieldNumber)
        ):
            return ~models.Q(
                **{"%s__exact" % reportrow.field_name: smart_str(data).strip()}
            )
        elif isinstance(reportrow, GridFieldChoice):
            t = smart_str(data).strip().lower()
            # Comparison also with the translated choices
            for c in reportrow.choices:
                if t == force_str(c[1]).lower():
                    return ~models.Q(**{"%s__iexact" % reportrow.field_name: c[0]})
            return ~models.Q(
                **{"%s__iexact" % reportrow.field_name: smart_str(data).strip()}
            )
        elif isinstance(reportrow, GridFieldDateTime):
            return ~models.Q(
                **{"%s__exact" % reportrow.field_name: parseLocalizedDateTime(data)}
            )
        elif isinstance(reportrow, GridFieldDate):
            return ~models.Q(
                **{"%s__exact" % reportrow.field_name: parseLocalizedDate(data)}
            )
        else:
            return ~models.Q(
                **{"%s__iexact" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_bn(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices only
            accepted = []
            t = smart_str(data).strip().lower()
            for c in reportrow.choices:
                if force_str(c[1]).lower().startswith(t):
                    accepted.append(c[0])
            return ~models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return ~models.Q(
                **{"%s__istartswith" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_en(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices only
            accepted = []
            t = smart_str(data).strip().lower()
            for c in reportrow.choices:
                if force_str(c[1]).lower().endswith(t):
                    accepted.append(c[0])
            return ~models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return ~models.Q(
                **{"%s__iendswith" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_nc(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(
            reportrow, (GridFieldCurrency, GridFieldInteger, GridFieldNumber)
        ):
            return ~models.Q(
                **{"%s__contains" % reportrow.field_name: smart_str(data).strip()}
            )
        elif isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices
            accepted = []
            t = data.strip().lower()
            for c in reportrow.choices:
                if t in force_str(c[1]).lower():
                    accepted.append(c[0])
            return ~models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return ~models.Q(
                **{"%s__icontains" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_ni(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison also with the translated choices
            accepted = []
            for f in smart_str(data).split(","):
                t = f.strip().lower()
                for c in reportrow.choices:
                    if t in (c[0].lower(), force_str(c[1]).lower()):
                        accepted.append(c[0])
            return ~models.Q(**{"%s__in" % reportrow.field_name: accepted})
        else:
            return ~models.Q(
                **{"%s__in" % reportrow.field_name: smart_str(data).strip().split(",")}
            )

    @staticmethod
    def _filter_in(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison also with the translated choices
            accepted = []
            for f in smart_str(data).split(","):
                t = f.strip().lower()
                for c in reportrow.choices:
                    if t in (c[0].lower(), force_str(c[1]).lower()):
                        accepted.append(c[0])
            return models.Q(**{"%s__in" % reportrow.field_name: accepted})
        else:
            return models.Q(
                **{"%s__in" % reportrow.field_name: smart_str(data).strip().split(",")}
            )

    @staticmethod
    def _filter_eq(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(
            reportrow, (GridFieldCurrency, GridFieldInteger, GridFieldNumber)
        ):
            return models.Q(
                **{"%s__exact" % reportrow.field_name: smart_str(data).strip()}
            )
        elif isinstance(reportrow, GridFieldChoice):
            t = smart_str(data).strip().lower()
            # Comparison with the translated choices only
            for c in reportrow.choices:
                if t == force_str(c[1]).lower():
                    return models.Q(**{"%s__iexact" % reportrow.field_name: c[0]})
            return models.Q(
                **{"%s__iexact" % reportrow.field_name: smart_str(data).strip()}
            )
        elif isinstance(reportrow, GridFieldDateTime):
            return models.Q(
                **{"%s__exact" % reportrow.field_name: parseLocalizedDateTime(data)}
            )
        elif isinstance(reportrow, GridFieldDate):
            return models.Q(
                **{"%s__exact" % reportrow.field_name: parseLocalizedDate(data)}
            )
        else:
            return models.Q(
                **{"%s__iexact" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_bw(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices only
            accepted = []
            t = data.strip().lower()
            for c in reportrow.choices:
                if force_str(c[1]).lower().startswith(t):
                    accepted.append(c[0])
            return models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return models.Q(
                **{"%s__istartswith" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_gt(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldDateTime):
            data = parseLocalizedDateTime(data)
        elif isinstance(reportrow, GridFieldDate):
            data = parseLocalizedDate(data)
        return models.Q(**{"%s__gt" % reportrow.field_name: smart_str(data).strip()})

    @staticmethod
    def _filter_gte(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldDateTime):
            data = parseLocalizedDateTime(data)
        elif isinstance(reportrow, GridFieldDate):
            data = parseLocalizedDate(data)
        return models.Q(**{"%s__gte" % reportrow.field_name: smart_str(data).strip()})

    @staticmethod
    def _filter_lt(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldDateTime):
            data = parseLocalizedDateTime(data)
        elif isinstance(reportrow, GridFieldDate):
            data = parseLocalizedDate(data)
        return models.Q(**{"%s__lt" % reportrow.field_name: smart_str(data).strip()})

    @staticmethod
    def _filter_lte(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldDateTime):
            data = parseLocalizedDateTime(data)
        elif isinstance(reportrow, GridFieldDate):
            data = parseLocalizedDate(data)
        return models.Q(**{"%s__lte" % reportrow.field_name: smart_str(data).strip()})

    @staticmethod
    def _filter_ew(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices
            accepted = []
            t = data.strip().lower()
            for c in reportrow.choices:
                if force_str(c[1]).lower().endswith(t):
                    accepted.append(c[0])
            return models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return models.Q(
                **{"%s__iendswith" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_cn(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if isinstance(reportrow, GridFieldChoice):
            # Comparison with the translated choices
            accepted = []
            t = data.strip().lower()
            for c in reportrow.choices:
                if t in force_str(c[1]).lower():
                    accepted.append(c[0])
            return models.Q(
                **{"%s__in" % reportrow.field_name: accepted or ["--dummy--"]}
            )
        else:
            return models.Q(
                **{"%s__icontains" % reportrow.field_name: smart_str(data).strip()}
            )

    @staticmethod
    def _filter_win(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        limit = date.today() + timedelta(int(float(smart_str(data))))
        return models.Q(**{"%s__lte" % reportrow.field_name: limit})

    @staticmethod
    def _filter_isnull(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        if data.lower() in ["0", "false", force_str(_("false"))]:
            return models.Q(**{"%s__isnull" % reportrow.field_name: False})
        else:
            return models.Q(**{"%s__isnull" % reportrow.field_name: True})

    @staticmethod
    def _filter_ico(query, reportrow, data, database=DEFAULT_DB_ALIAS):
        # rebuild hierarchy
        if issubclass(reportrow.model, HierarchyModel):
            reportrow.model.rebuildHierarchy(database)
        else:
            raise Exception("ico filter can only be used on hierarchical models")

        # get parent object
        try:
            o = reportrow.model.objects.using(database).get(name__iexact=data)
            parentExists = True
        except Exception:
            parentExists = False

        prefix = not (reportrow.name == "name")

        return models.Q(
            **{
                "%slft__gte"
                % ("%s__" % reportrow.model.__name__.lower() if prefix else ""): (
                    o.lft if parentExists else -1
                ),
                "%slft__lte"
                % ("%s__" % reportrow.model.__name__.lower() if prefix else ""): (
                    o.rght if parentExists else -1
                ),
            }
        )

    _filter_map_jqgrid_django = {
        # jqgrid op: (django_lookup, use_exclude, use_extra_where)
        "ne": _filter_ne.__func__,
        "bn": _filter_bn.__func__,
        "en": _filter_en.__func__,
        "nc": _filter_nc.__func__,
        "ni": _filter_ni.__func__,
        "in": _filter_in.__func__,
        "eq": _filter_eq.__func__,
        "bw": _filter_bw.__func__,
        "gt": _filter_gt.__func__,
        "ge": _filter_gte.__func__,
        "lt": _filter_lt.__func__,
        "le": _filter_lte.__func__,
        "ew": _filter_ew.__func__,
        "cn": _filter_cn.__func__,
        "win": _filter_win.__func__,
        "ico": _filter_ico.__func__,
        "isnull": _filter_isnull.__func__,
    }

    _filter_map_django_jqgrid = {
        # django lookup: jqgrid op
        "in": "in",
        "exact": "eq",
        "startswith": "bw",
        "iexact": "eq",
        "istartswith": "bw",
        "gt": "gt",
        "gte": "ge",
        "lt": "lt",
        "lte": "le",
        "endswith": "ew",
        "contains": "cn",
        "iendswith": "ew",
        "icontains": "cn",
        "ico": "ico",
        "isnull": "isnull",
        # 'win' exist in jqgrid, but not in django
        # 'ne' exist in jqgrid, but not in django
    }

    @classmethod
    def getQueryString(cls, request):
        # Django-style filtering (which uses URL parameters) are converted to a jqgrid filter expression
        filters = {"groupOp": "AND", "rules": []}
        for i, j in request.GET.items():
            for r in request.rows:
                if i == r.name or (
                    r.field_name
                    and (i == r.field_name or i.startswith(r.field_name + "__"))
                ):
                    operator = (
                        (i == r.field_name or i == r.name)
                        and "exact"
                        or i[i.rfind("__") + 2 :]
                    )
                    try:
                        filters["rules"].append(
                            {
                                "field": r.name,
                                "op": cls._filter_map_django_jqgrid[operator],
                                "data": unquote(j),
                            }
                        )
                    except Exception:
                        pass  # Ignore invalid operators
        return json.dumps(filters) if filters["rules"] else None

    @classmethod
    def _get_q_filter(cls, request, filterdata):
        q_filters = []
        for rule in filterdata.get("rules", []):
            try:
                op, field, data = rule["op"], rule["field"], rule["data"]
                reportrow = cls._getRowByName(request, field)
                if data == "":
                    # No filter value specified, which makes the filter invalid
                    continue
                else:
                    q_filters.append(
                        cls._filter_map_jqgrid_django[op](
                            q_filters, reportrow, data, request.database
                        )
                    )
            except Exception as e:
                pass  # Silently ignore invalid filters
        for group in filterdata.get("groups", []):
            try:
                z = cls._get_q_filter(request, group)
                if z:
                    q_filters.append(z)
            except Exception:
                pass  # Silently ignore invalid groups
        if len(q_filters) == 0:
            return None
        elif filterdata["groupOp"].upper() == "OR":
            return functools.reduce(operator.ior, q_filters)
        else:
            return functools.reduce(operator.iand, q_filters)

    @classmethod
    def filter_items(cls, request, items, plus_django_style=True):
        # Jqgrid-style advanced filtering
        filters = None
        _filters = request.GET.get("filters")
        if _filters:
            # Validate complex search JSON data
            try:
                filters = json.loads(_filters)
            except ValueError:
                filters = None

        # Single field searching, which is currently not used
        if request.GET.get("_search") == "true" and not filters:
            field = request.GET.get("searchField")
            op = request.GET.get("searchOper")
            data = request.GET.get("searchString")
            if all([field, op, data]):
                filters = {
                    "groupOp": "AND",
                    "rules": [{"op": op, "field": field, "data": data}],
                }

        if filters:
            z = cls._get_q_filter(request, filters)
            if z:
                return items.filter(z)
            else:
                return items

        # Django-style filtering, using URL parameters
        if plus_django_style:
            for i, j in request.GET.items():
                for r in request.rows:
                    if r.name and (
                        i == r.field_name or i.startswith(r.field_name + "__")
                    ):
                        try:
                            items = items.filter(**{i: unquote(j)})
                        except Exception:
                            pass  # silently ignore invalid filters
        return items


class GridPivot(GridReport):
    # Cross definitions.
    # Possible attributes for a cross field are:
    #   - title:
    #     Name of the cross that is displayed to the user.
    #     It defaults to the name of the field.
    #   - editable:
    #     True when the field is editable in the page.
    #     The default value is false.
    #   - initially_hidden:
    #     True when this field should be hidden by default.
    crosses = ()

    template = "admin/base_site_gridpivot.html"

    hasTimeBuckets = True

    editable = False

    multiselect = False

    @classmethod
    def _render_cross(cls, request):
        result = []
        for i in request.crosses:
            m = {"key": i[0]}
            for key, value in i[1].items():
                if callable(value):
                    if key == "title":
                        m["name"] = capfirst(force_str(value(request)))
                    else:
                        tmp = value(request)
                        if isinstance(tmp, list):
                            m[key] = [force_str(v, strings_only=True) for v in tmp]
                        else:
                            m[key] = force_str(tmp, strings_only=True)
                else:
                    if key == "title":
                        m["name"] = capfirst(force_str(value))
                    elif isinstance(value, list):
                        m[key] = [force_str(v, strings_only=True) for v in value]
                    else:
                        m[key] = force_str(value, strings_only=True)
            if "editable" not in m:
                m["editable"] = False
            result.append(json.dumps(m))
        return ",\n".join(result)

    @classmethod
    def _render_colmodel(
        cls, request, is_popup=False, prefs=None, mode="graph", *args, **kwargs
    ):
        if prefs and "rows" in prefs:
            # Validate the preferences to 1) map from name to index, 2) assure all rows
            # are included, 3) ignore non-existing fields
            prefrows = prefs["rows"]
            defaultrows = {request.rows[i].name: i for i in range(len(request.rows))}
            rows = []
            for r in prefrows:
                try:
                    idx = int(r[0])
                    defaultrows.pop(request.rows[idx].name, None)
                    rows.append(r)
                except (ValueError, IndexError):
                    if r[0] in defaultrows:
                        rows.append((defaultrows[r[0]], r[1], r[2]))
                        defaultrows.pop(r[0], None)
            for r, idx in defaultrows.items():
                rows.append(
                    (
                        idx,
                        request.rows[idx].hidden or request.rows[idx].initially_hidden,
                        request.rows[idx].width,
                    )
                )
        else:
            # Default configuration
            rows = [
                (
                    i,
                    request.rows[i].initially_hidden or request.rows[i].hidden,
                    request.rows[i].width,
                )
                for i in range(len(request.rows))
            ]

        result = []
        if is_popup:
            result.append(
                '{"name":"select","label":gettext("Select"),"width":75,"align":"center","sortable":false,"search":false,"fixed":true}'
            )
        count = -1
        for index, hidden, width in rows:
            try:
                result.append(
                    '{%s,"width":%s,"counter":%d,"frozen":true%s,"hidden":%s,"fixed":true}'
                    % (
                        request.rows[index],
                        width,
                        index,
                        is_popup and ',"popup":true' or "",
                        hidden and "true" or "false",
                    )
                )
                count += 1
            except IndexError:
                pass
        return ",\n".join(result)

    @classmethod
    def count_query(cls, request, *args, **kwargs):
        if not hasattr(request, "basequery"):
            if callable(cls.basequeryset):
                request.basequery = cls.basequeryset(request, *args, **kwargs)
            else:
                request.basequery = cls.basequeryset
            if args and args[0] and not cls.new_arg_logic:
                request.basequery = request.basequery.filter(pk__exact=args[0])
        return (
            cls.filter_items(request, request.basequery).using(request.database).count()
        )

    @classmethod
    def data_query(cls, request, *args, page=None, fields=None, **kwargs):
        if not fields:
            raise Exception("No fields for pivot report")
        if not hasattr(request, "basequery"):
            if callable(cls.basequeryset):
                request.basequery = cls.basequeryset(request, *args, **kwargs)
            else:
                request.basequery = cls.basequeryset
            if args and args[0] and not cls.new_arg_logic:
                request.basequery = request.basequery.filter(pk__exact=args[0])
        if page:
            cnt = (page - 1) * request.pagesize + 1
            return cls.query(
                request,
                cls._apply_sort(
                    request, cls.filter_items(request, request.basequery)
                ).using(request.database)[cnt - 1 : cnt + request.pagesize],
                sortsql=cls._apply_sort_index(request),
            )
        else:
            return cls.query(
                request,
                cls._apply_sort(
                    request, cls.filter_items(request, request.basequery)
                ).using(request.database),
                sortsql=cls._apply_sort_index(request),
            )

    @classmethod
    def _generate_json_data(cls, request, *args, **kwargs):
        # Prepare the query
        request.prefs = request.user.getPreference(
            cls.getKey(request, *args, **kwargs), database=request.database
        )
        recs = cls.count_query(request, *args, **kwargs)
        page = "page" in request.GET and int(request.GET["page"]) or 1
        if "rows" in request.GET:
            request.pagesize = int(request.GET["rows"])
        total_pages = math.ceil(float(recs) / request.pagesize)
        if page > total_pages:
            page = total_pages
        if page < 1:
            page = 1

        # Generate header of the output
        yield '{"total":%d,\n' % total_pages
        yield '"page":%d,\n' % page
        yield '"records":%d,\n' % recs
        yield '"rows":[\n'

        # Generate output
        currentkey = None
        r = []
        fields = [i.field_name for i in request.rows if i.field_name]
        for i in cls.data_query(request, *args, page=page, fields=fields, **kwargs):
            # We use the first field in the output to recognize new rows.
            if currentkey != i[request.rows[0].name]:
                # New line
                if currentkey:
                    yield "".join(r)
                    r = ["},\n{"]
                else:
                    r = ["{"]
                currentkey = i[request.rows[0].name]
                first2 = True
                for f in request.rows:
                    try:
                        s = cls._getJSONValue(i[f.name], field=f, request=request)
                        if first2:
                            r.append('"%s":%s' % (f.name, s))
                            first2 = False
                        elif i[f.name] is not None:
                            r.append(', "%s":%s' % (f.name, s))
                    except Exception:
                        pass
            r.append(', "%s":[' % i["bucket"])
            first2 = True
            for f in request.crosses:
                if i[f[0]] is None:
                    if first2:
                        r.append("null")
                        first2 = False
                    else:
                        r.append(",null")
                else:
                    if first2:
                        r.append("%s" % i[f[0]])
                        first2 = False
                    else:
                        r.append(",%s" % i[f[0]])
            r.append("]")
        r.append("}")
        r.append("\n]}\n")
        yield "".join(r)

    @classmethod
    def _generate_csv_data(cls, request, scenario_list, *args, **kwargs):
        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue(
                "excel_duration_in_days", request.database, "false"
            ).lower()
            == "true"
        )

        sf = StringIO()
        decimal_separator = get_format("DECIMAL_SEPARATOR", request.LANGUAGE_CODE, True)
        if decimal_separator == ",":
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=";")
        else:
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=",")
        if translation.get_language() != request.LANGUAGE_CODE:
            translation.activate(request.LANGUAGE_CODE)
        listformat = request.GET.get("format", "csvlist") == "csvlist"

        # Write a Unicode Byte Order Mark header, aka BOM (Excel needs it to open UTF-8 file properly)
        yield cls.getBOM(settings.CSV_CHARSET)

        # Pick up the preferences
        if not hasattr(request, "prefs"):
            request.prefs = request.user.getPreference(
                cls.getKey(request, *args, **kwargs), database=request.database
            )

        allColumns = request.GET.get("allcolumns", False)

        if request.prefs and "rows" in request.prefs and not allColumns:
            myrows = [
                request.rows[f[0]]
                for f in cls._validate_rows(request, request.prefs["rows"])
                if not f[1]
            ]
        else:
            myrows = [
                f
                for f in request.rows
                if f.name and not f.hidden and not f.initially_hidden
            ]
        if request.prefs and "crosses" in request.prefs and not allColumns:
            mycrosses = [
                request.crosses[f]
                for f in cls._validate_crosses(request, request.prefs["crosses"])
            ]
        else:
            mycrosses = [f for f in request.crosses if f[1].get("visible", True)]

        # Write a header row
        fields = [
            force_str(f.title, encoding=settings.CSV_CHARSET, errors="ignore").title()
            for f in myrows
            if f.name
        ]
        if listformat:
            fields.extend(
                [
                    capfirst(
                        force_str(
                            _("bucket"), encoding=settings.CSV_CHARSET, errors="ignore"
                        )
                    )
                ]
            )
            fields.extend(
                [
                    capfirst(
                        force_str(
                            _(
                                (
                                    f[1]["title"](request)
                                    if callable(f[1]["title"])
                                    else f[1]["title"]
                                )
                                if "title" in f[1]
                                else f[0]
                            ),
                            encoding=settings.CSV_CHARSET,
                            errors="ignore",
                        )
                    )
                    for f in mycrosses
                ]
            )
        else:
            fields.extend(
                [
                    capfirst(
                        force_str(
                            _("data field"),
                            encoding=settings.CSV_CHARSET,
                            errors="ignore",
                        )
                    )
                ]
            )
            fields.extend(
                [
                    force_str(b["name"], encoding=settings.CSV_CHARSET, errors="ignore")
                    for b in request.report_bucketlist
                ]
            )
        if len(scenario_list) > 1:
            fields.insert(0, _("scenario"))
        writer.writerow(fields)
        yield sf.getvalue()

        # Write the report content
        orginal_database = request.database
        try:
            for scenario in scenario_list:
                request.database = scenario
                if hasattr(request, "query"):
                    delattr(request, "query")

                query = cls.data_query(request, *args, fields=fields, **kwargs)

                if listformat:
                    for row in query:
                        # Clear the return string buffer
                        sf.seek(0)
                        sf.truncate(0)
                        # Data for rows
                        if hasattr(row, "__getitem__"):
                            fields = [
                                cls._getCSVValue(
                                    row[f.name],
                                    field=f,
                                    request=request,
                                    decimal_separator=decimal_separator,
                                    excel_duration_in_days=excel_duration_in_days,
                                )
                                for f in myrows
                                if f.name
                            ]
                            fields.extend(
                                [
                                    force_str(
                                        row["bucket"],
                                        encoding=settings.CSV_CHARSET,
                                        errors="ignore",
                                    )
                                ]
                            )
                            fields.extend(
                                [
                                    (
                                        force_str(
                                            cls._localize(row[f[0]], decimal_separator),
                                            encoding=settings.CSV_CHARSET,
                                            errors="ignore",
                                        )
                                        if row[f[0]] is not None
                                        else ""
                                    )
                                    for f in mycrosses
                                ]
                            )
                        else:
                            fields = [
                                cls._getCSVValue(
                                    getattr(row, f.name),
                                    field=f,
                                    request=request,
                                    decimal_separator=decimal_separator,
                                    excel_duration_in_days=excel_duration_in_days,
                                )
                                for f in myrows
                                if f.name
                            ]
                            fields.extend(
                                [
                                    force_str(
                                        getattr(row, "bucket"),
                                        encoding=settings.CSV_CHARSET,
                                        errors="ignore",
                                    )
                                ]
                            )
                            fields.extend(
                                [
                                    (
                                        force_str(
                                            cls._localize(
                                                getattr(row, f[0]), decimal_separator
                                            ),
                                            encoding=settings.CSV_CHARSET,
                                            errors="ignore",
                                        )
                                        if getattr(row, f[0]) is not None
                                        else ""
                                    )
                                    for f in mycrosses
                                ]
                            )
                        # Return string
                        if len(scenario_list) > 1:
                            fields.insert(0, scenario_list[scenario])
                        writer.writerow(fields)
                        yield sf.getvalue()
                else:
                    currentkey = None
                    row_of_buckets = None
                    for row in query:
                        # We use the first field in the output to recognize new rows.
                        if not currentkey:
                            currentkey = row[request.rows[0].name]
                            row_of_buckets = [row]
                        elif currentkey == row[request.rows[0].name]:
                            row_of_buckets.append(row)
                        else:
                            # Write an entity
                            for cross in mycrosses:
                                # Clear the return string buffer
                                sf.seek(0)
                                sf.truncate(0)
                                fields = [
                                    cls._getCSVValue(
                                        row_of_buckets[0][s.name],
                                        field=s,
                                        request=request,
                                        decimal_separator=decimal_separator,
                                        excel_duration_in_days=excel_duration_in_days,
                                    )
                                    for s in myrows
                                    if s.name
                                ]
                                fields.extend(
                                    [
                                        force_str(
                                            capfirst(
                                                _(
                                                    (
                                                        cross[1]["title"](request)
                                                        if callable(cross[1]["title"])
                                                        else cross[1]["title"]
                                                    )
                                                    if "title" in cross[1]
                                                    else cross[0]
                                                )
                                            ),
                                            encoding=settings.CSV_CHARSET,
                                            errors="ignore",
                                        )
                                    ]
                                )
                                fields.extend(
                                    [
                                        (
                                            force_str(
                                                cls._localize(
                                                    bucket[cross[0]], decimal_separator
                                                ),
                                                encoding=settings.CSV_CHARSET,
                                                errors="ignore",
                                            )
                                            if bucket[cross[0]] is not None
                                            else ""
                                        )
                                        for bucket in row_of_buckets
                                    ]
                                )
                                # Return string
                                if len(scenario_list) > 1:
                                    fields.insert(0, scenario_list[scenario])
                                writer.writerow(fields)
                                yield sf.getvalue()
                            currentkey = row[request.rows[0].name]
                            row_of_buckets = [row]
                    # Write the last entity
                    if row_of_buckets:
                        for cross in mycrosses:
                            # Clear the return string buffer
                            sf.seek(0)
                            sf.truncate(0)
                            fields = [
                                cls._getCSVValue(
                                    row_of_buckets[0][s.name],
                                    field=s,
                                    request=request,
                                    decimal_separator=decimal_separator,
                                    excel_duration_in_days=excel_duration_in_days,
                                )
                                for s in myrows
                                if s.name
                            ]
                            fields.extend(
                                [
                                    force_str(
                                        capfirst(
                                            _(
                                                (
                                                    cross[1]["title"](request)
                                                    if callable(cross[1]["title"])
                                                    else cross[1]["title"]
                                                )
                                                if "title" in cross[1]
                                                else cross[0]
                                            )
                                        ),
                                        encoding=settings.CSV_CHARSET,
                                        errors="ignore",
                                    )
                                ]
                            )
                            fields.extend(
                                [
                                    force_str(
                                        cls._localize(
                                            bucket[cross[0]], decimal_separator
                                        ),
                                        encoding=settings.CSV_CHARSET,
                                        errors="ignore",
                                    )
                                    for bucket in row_of_buckets
                                ]
                            )
                            # Return string
                            if len(scenario_list) > 1:
                                fields.insert(0, scenario_list[scenario])
                            writer.writerow(fields)
                            yield sf.getvalue()
        finally:
            request.database = orginal_database

    @classmethod
    def _generate_spreadsheet_data(
        cls, request, scenario_list, output, *args, **kwargs
    ):
        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue(
                "excel_duration_in_days", request.database, "false"
            ).lower()
            == "true"
        )

        # Create a workbook
        wb = Workbook(write_only=True)
        if callable(cls.title):
            title = cls.title(request, *args, **kwargs)
        elif hasattr(cls, "title"):
            title = cls.title
        else:
            title = cls.model._meta.verbose_name_plural if cls.model else cls.title
        wb.properties.creator = "frepple %s" % __version__
        ws = wb.create_sheet(title=cls._sanitize_excel_tab(title))

        # Create a named style for the header row
        headerstyle = NamedStyle(name="headerstyle")
        headerstyle.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
        wb.add_named_style(headerstyle)
        readlonlyheaderstyle = NamedStyle(name="readlonlyheaderstyle")
        readlonlyheaderstyle.fill = PatternFill(fill_type="solid", fgColor="d0ebfb")
        wb.add_named_style(readlonlyheaderstyle)

        # Pick up the preferences
        listformat = request.GET.get("format", "spreadsheetlist") == "spreadsheetlist"
        if not hasattr(request, "prefs"):
            request.prefs = request.user.getPreference(
                cls.getKey(request, *args, **kwargs), database=request.database
            )
        rows = request.GET.get("selected_rows", None)
        if rows:
            rows = rows.split(",")
            myrows = [f for f in request.rows if f.name in rows]
            myrows.sort(key=lambda x: rows.index(x.name))
        elif request.prefs and "rows" in request.prefs:
            myrows = [
                request.rows[f[0]]
                for f in cls._validate_rows(request, request.prefs["rows"])
                if not f[1]
            ]
        else:
            myrows = [
                f
                for f in request.rows
                if f.name and not f.initially_hidden and not f.hidden
            ]
        crosses = request.GET.get("selected_crosses", None)
        if crosses:
            crosses = crosses.split(",")
            mycrosses = [f for f in request.crosses if f[0] in crosses]
            mycrosses.sort(key=lambda x: crosses.index(x[0]))
        elif request.prefs and "crosses" in request.prefs:
            mycrosses = [
                request.crosses[f]
                for f in cls._validate_crosses(request, request.prefs["crosses"])
            ]
        else:
            mycrosses = [f for f in request.crosses if f[1].get("visible", True)]

        # Create custom named styles
        backgrounds = set()
        for r in myrows:
            if r.background_header:
                backgrounds.add(r.background_header)
            if r.background_cell:
                backgrounds.add(r.background_cell)
        for r in mycrosses:
            if r[1].get("background_header", None):
                backgrounds.add(r[1].get("background_header", None))
            if r[1].get("background_cell", None):
                backgrounds.add(r[1].get("background_cell", None))
        for bg in backgrounds:
            style = NamedStyle(name=bg)
            style.fill = PatternFill(fill_type="solid", fgColor=bg)
            wb.add_named_style(style)

        # Write a header row
        fields = []
        comment = None
        for f in myrows:
            if f.name:
                cell = WriteOnlyCell(ws, value=force_str(f.title).title())
                if f.editable or f.key:
                    cell.style = f.background_header or "headerstyle"
                    fname = getattr(f, "field_name", f.name)
                    if (
                        not f.key
                        and f.formatter == "detail"
                        and fname.endswith("__name")
                    ):
                        cell.comment = CellComment(
                            force_str(
                                _("Values in this field must exist in the %s table")
                                % force_str(_(fname[:-6]))
                            ),
                            "Author",
                        )
                    elif isinstance(f, GridFieldChoice):
                        cell.comment = CellComment(
                            force_str(
                                _("Accepted values are: %s")
                                % ", ".join([c[0] for c in f.choices])
                            ),
                            "Author",
                        )
                else:
                    cell.style = f.background_header or "readlonlyheaderstyle"
                    if not comment:
                        comment = CellComment(
                            force_str(_("Read only")), "Author", height=20, width=80
                        )
                    cell.comment = comment
                fields.append(cell)
        if listformat:
            cell = WriteOnlyCell(ws, value=capfirst(force_str(_("bucket"))))
            if f.editable or f.key:
                cell.style = f.background_header or "headerstyle"
                fname = getattr(f, "field_name", f.name)
                if not f.key and f.formatter == "detail" and fname.endswith("__name"):
                    cell.comment = CellComment(
                        force_str(
                            _("Values in this field must exist in the %s table")
                            % force_str(_(fname[:-6]))
                        ),
                        "Author",
                    )
                elif isinstance(f, GridFieldChoice):
                    cell.comment = CellComment(
                        force_str(
                            _("Accepted values are: %s")
                            % ", ".join([c[0] for c in f.choices])
                        ),
                        "Author",
                    )
            else:
                cell.style = f.background_header or "readlonlyheaderstyle"
                if not comment:
                    comment = CellComment(
                        force_str(_("Read only")), "Author", height=20, width=80
                    )
                cell.comment = comment
            fields.append(cell)
            for f in mycrosses:
                cell = WriteOnlyCell(
                    ws,
                    value=capfirst(
                        force_str(
                            _(
                                (
                                    f[1]["title"](request)
                                    if callable(f[1]["title"])
                                    else f[1]["title"]
                                )
                                if "title" in f[1]
                                else f[0]
                            )
                        )
                    ),
                )
                if f[1].get("editable", False):
                    cell.style = f[1].get("background_header", None) or "headerstyle"
                else:
                    cell.style = (
                        f[1].get("background_header", None) or "readlonlyheaderstyle"
                    )
                    if not comment:
                        comment = CellComment(
                            force_str(_("Read only")), "Author", height=20, width=80
                        )
                    cell.comment = comment
                fields.append(cell)
        else:
            cell = WriteOnlyCell(ws, value=capfirst(force_str(_("data field"))))
            cell.style = "readlonlyheaderstyle"
            fields.append(cell)
            for b in request.report_bucketlist:
                cell = WriteOnlyCell(ws, value=str(b["name"]))
                cell.style = "readlonlyheaderstyle"
                fields.append(cell)

        if len(scenario_list) > 1:
            cell = WriteOnlyCell(ws, value=capfirst(force_str(_("scenario"))))
            cell.style = "readlonlyheaderstyle"
            fields.insert(0, cell)

        ws.append(fields)

        # Add an auto-filter to the table
        ws.auto_filter.ref = "A1:%s1048576" % get_column_letter(len(fields))

        # Write the report content
        original_database = request.database
        try:
            for scenario in scenario_list:
                request.database = scenario
                if hasattr(request, "query"):
                    delattr(request, "query")

                query = cls.data_query(request, *args, fields=fields, **kwargs)

                if listformat:
                    for row in query:
                        # Append a row
                        fields = []
                        if hasattr(row, "__getitem__"):
                            for f in myrows:
                                if f.name:
                                    cell = WriteOnlyCell(
                                        ws,
                                        value=_getCellValue(
                                            row[f.name],
                                            field=f,
                                            request=request,
                                            excel_duration_in_days=excel_duration_in_days,
                                        ),
                                    )
                                    if f.background_cell:
                                        cell.style = f.background_cell
                                    fields.append(cell)
                            fields.append(_getCellValue(row["bucket"]))
                            for f in mycrosses:
                                cell = WriteOnlyCell(ws, value=_getCellValue(row[f[0]]))
                                if f[1].get("background_cell"):
                                    cell.style = f[1].get("background_cell")
                                fields.append(cell)
                        else:
                            for f in myrows:
                                if f.name:
                                    cell = WriteOnlyCell(
                                        ws,
                                        value=_getCellValue(
                                            getattr(row, f.name),
                                            field=f,
                                            request=request,
                                            excel_duration_in_days=excel_duration_in_days,
                                        ),
                                    )
                                    if f.background_cell:
                                        cell.style = f.background_cell
                                    fields.append(cell)
                            fields.append(_getCellValue(getattr(row, "bucket")))
                            for f in mycrosses:
                                cell = WriteOnlyCell(
                                    ws, value=_getCellValue(getattr(row, f[0]))
                                )
                                if f[1].get("background_cell"):
                                    cell.style = f[1].get("background_cell")
                                fields.append(cell)
                        if len(scenario_list) > 1:
                            fields.insert(0, scenario_list[scenario])
                        ws.append(fields)
                else:
                    currentkey = None
                    row_of_buckets = None
                    for row in query:
                        # We use the first field in the output to recognize new rows.
                        if not currentkey:
                            currentkey = row[request.rows[0].name]
                            row_of_buckets = [row]
                        elif currentkey == row[request.rows[0].name]:
                            row_of_buckets.append(row)
                        else:
                            # Write a row
                            for cross in mycrosses:
                                if not cross[1].get("visible", True):
                                    continue
                                fields = []
                                for s in myrows:
                                    if s.name:
                                        cell = WriteOnlyCell(
                                            ws,
                                            value=_getCellValue(
                                                row_of_buckets[0][s.name],
                                                field=s,
                                                request=request,
                                                excel_duration_in_days=excel_duration_in_days,
                                            ),
                                        )
                                        if s.background_cell:
                                            cell.style = s.background_cell
                                        fields.append(cell)
                                cell = WriteOnlyCell(
                                    ws,
                                    value=_getCellValue(
                                        (
                                            (
                                                capfirst(
                                                    cross[1]["title"](request)
                                                    if callable(cross[1]["title"])
                                                    else cross[1]["title"]
                                                )
                                            )
                                            if "title" in cross[1]
                                            else capfirst(cross[0])
                                        ),
                                        excel_duration_in_days=excel_duration_in_days,
                                    ),
                                )
                                if cross[1].get("background_header"):
                                    cell.style = cross[1].get("background_header")
                                fields.append(cell)
                                for bucket in row_of_buckets:
                                    cell = WriteOnlyCell(
                                        ws, value=_getCellValue(bucket[cross[0]])
                                    )
                                    if cross[1].get("background_cell"):
                                        cell.style = cross[1].get("background_cell")
                                    fields.append(cell)
                                if len(scenario_list) > 1:
                                    fields.insert(0, scenario_list[scenario])
                                ws.append(fields)
                            currentkey = row[request.rows[0].name]
                            row_of_buckets = [row]
                    # Write the last row
                    if row_of_buckets:
                        for cross in mycrosses:
                            if not cross[1].get("visible", True):
                                continue
                            fields = []
                            for s in myrows:
                                if s.name:
                                    cell = WriteOnlyCell(
                                        ws,
                                        value=_getCellValue(
                                            row_of_buckets[0][s.name],
                                            field=s,
                                            request=request,
                                            excel_duration_in_days=excel_duration_in_days,
                                        ),
                                    )
                                    if s.background_cell:
                                        cell.style = s.background_cell
                                    fields.append(cell)
                            cell = WriteOnlyCell(
                                ws,
                                value=_getCellValue(
                                    (
                                        capfirst(
                                            cross[1]["title"](request)
                                            if callable(cross[1]["title"])
                                            else cross[1]["title"]
                                        )
                                    )
                                    if "title" in cross[1]
                                    else capfirst(cross[0])
                                ),
                            )
                            if cross[1].get("background_header"):
                                cell.style = cross[1].get("background_header")
                            fields.append(cell)
                            for bucket in row_of_buckets:
                                cell = WriteOnlyCell(
                                    ws, value=_getCellValue(bucket[cross[0]])
                                )
                                if cross[1].get("background_cell"):
                                    cell.style = cross[1].get("background_cell")
                                fields.append(cell)
                            if len(scenario_list) > 1:
                                fields.insert(0, scenario_list[scenario])
                            ws.append(fields)
        finally:
            request.database = original_database

        # Write the spreadsheet
        wb.save(output)


numericTypes = (Decimal, float, int)


def _buildMaskedNames(model, exportConfig):
    """
    Build a map with anonymous names for a model, and store it in the exportConfiguration.
    """
    modelname = model._meta.model_name
    if modelname in exportConfig:
        return
    exportConfig[modelname] = {}
    if issubclass(model, HierarchyModel):
        keys = (
            model.objects.only("pk").order_by("lvl", "pk").values_list("pk", flat=True)
        )
    else:
        keys = model.objects.only("pk").order_by("pk").values_list("pk", flat=True)
    idx = 1
    for key in keys:
        exportConfig[modelname][key] = "%s %07d" % (modelname, idx)
        idx += 1


def _parseSeconds(data, excel_duration_in_days=False):
    """
    Formats a number of seconds into format [D d] HH:MM:SS.XXXX
    if excel_duration_in_days is true:
        - duration that are less than a day will be returned in hh:mm:ss format
        - duration that are more than a day will be returned in float.
          E.g: 3.25 is 3 days and 6 hours
    """
    total_seconds = data.total_seconds()
    hours = math.floor(total_seconds / 3600)
    if hours >= 24:
        if not excel_duration_in_days:
            days = math.floor(hours / 24)
            total_seconds -= days * 86400
            hours = hours - days * 24
            if not total_seconds:
                return "%s d" % days
        else:
            days = total_seconds / 86400.0
            if days.is_integer():
                return int(days)
            else:
                return round(days, 8)
    else:
        days = 0

    if excel_duration_in_days:
        return data

    minutes = math.floor((total_seconds - hours * 3600) / 60)
    seconds = math.floor(total_seconds - hours * 3600 - minutes * 60)
    remainder = total_seconds - 3600 * hours - 60 * minutes - seconds
    return "%s%02d:%02d:%02d%s" % (
        "%s d " % days if days else "",
        hours,
        minutes,
        seconds,
        (".%s" % str(round(remainder, 8))[2:]) if remainder > 0 else "",
    )


def _getCellValue(
    data, field=None, exportConfig=None, request=None, excel_duration_in_days=False
):
    if data is None:
        return ""
    elif isinstance(data, datetime):
        if (
            field
            and request
            and isinstance(field, (GridFieldLastModified, GridFieldLocalDateTime))
        ):
            if not hasattr(request, "tzoffset"):
                request.tzoffset = GridReport.getTimezoneOffset(request)
            return data + request.tzoffset
        else:
            return data
    elif isinstance(data, numericTypes) or isinstance(data, date):
        return data
    elif isinstance(data, timedelta):
        return _parseSeconds(data, excel_duration_in_days)
    elif isinstance(data, time):
        return data.isoformat()
    elif not exportConfig or not exportConfig.get("anonymous", False):
        return str(data)
    else:
        if field.primary_key and not isinstance(field, AutoField):
            model = field.model
        elif isinstance(field, RelatedField):
            model = field.related_model
        else:
            return str(data)
        if not getattr(model, "obfuscate", True):
            return str(data)
        modelname = model._meta.model_name
        if modelname not in exportConfig:
            _buildMaskedNames(model, exportConfig)
        # Return the mapped value
        return exportConfig[modelname].get(data, "unknown")


def sizeof_fmt(num):
    """
    Function to convert from bytes to human readable format.
    """
    if num == None:
        return ""
    for unit in ["", "K", "M", "G", "T", "P", "E", "Z"]:
        if abs(num) < 1024.0:
            return "%3.0f %sB" % (num, unit)
        num /= 1024.0
    return "%.0f %sB" % (num, "Yi")
