#
# Copyright (C) 2011-2016 by frePPLe bv
#
# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE
# LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION
# OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION
# WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#

import codecs
from datetime import datetime
from time import localtime, strftime
import csv
import gzip
from openpyxl import load_workbook
import os
import logging

from django.conf import settings
from django.contrib.auth import get_permission_codename
from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import connections, DEFAULT_DB_ALIAS, transaction
from django.template.loader import render_to_string
from django.utils import translation
from django.utils.formats import get_format
from django.utils.translation import gettext_lazy as _

from freppledb.execute.models import Task
from freppledb.common.middleware import _thread_locals
from freppledb.common.report import GridReport, matchesModelName, sizeof_fmt
from freppledb import __version__
from freppledb.common.dataload import parseCSVdata, parseExcelWorksheet
from freppledb.common.models import User, NotificationFactory, Parameter
from freppledb.common.report import EXCLUDE_FROM_BULK_OPERATIONS, create_connection
from freppledb.common.utils import getStorageUsage, get_databases

logger = logging.getLogger(__name__)


def timesince(st):
    return str(datetime.now() - st).split(".")[0]


class Command(BaseCommand):
    help = """
    Loads CSV files from the configured FILEUPLOADFOLDER folder into the frePPLe database.
    The data files should have the extension .csv or .csv.gz, and the file name should
    start with the name of the data model.
    """

    requires_system_checks = []

    def add_arguments(self, parser):
        parser.add_argument("--user", help="User running the command")
        parser.add_argument(
            "--database",
            default=DEFAULT_DB_ALIAS,
            help="Nominates a specific database to load the data into",
        )
        parser.add_argument(
            "--task",
            type=int,
            help="Task identifier (generated automatically if not provided)",
        )

    def get_version(self):
        return __version__

    def handle(self, **options):
        # Pick up the options
        now = datetime.now()
        self.database = options["database"]
        if self.database not in get_databases():
            raise CommandError("No database settings known for '%s'" % self.database)
        if options["user"]:
            try:
                self.user = (
                    User.objects.all()
                    .using(self.database)
                    .get(username=options["user"])
                )
            except Exception:
                raise CommandError("User '%s' not found" % options["user"])
        else:
            self.user = None
        timestamp = now.strftime("%Y%m%d%H%M%S")
        if self.database == DEFAULT_DB_ALIAS:
            logfile = "importfromfolder-%s.log" % timestamp
        else:
            logfile = "importfromfolder_%s-%s.log" % (self.database, timestamp)

        try:
            handler = logging.FileHandler(
                os.path.join(settings.FREPPLE_LOGDIR, logfile), encoding="utf-8"
            )
            logger.addHandler(handler)
            logger.propagate = False
        except Exception as e:
            print("%s Failed to open logfile %s: %s" % (datetime.now(), logfile, e))

        task = None
        errors = [0, 0]
        old_thread_locals = getattr(_thread_locals, "database", None)
        startofall = datetime.now()
        try:
            setattr(_thread_locals, "database", self.database)
            # Initialize the task
            if options["task"]:
                try:
                    task = (
                        Task.objects.all().using(self.database).get(pk=options["task"])
                    )
                except Exception:
                    raise CommandError("Task identifier not found")
                if (
                    task.started
                    or task.finished
                    or task.status != "Waiting"
                    or task.name not in ("frepple_importfromfolder", "importfromfolder")
                ):
                    raise CommandError("Invalid task identifier")
                task.status = "0%"
                task.started = now
                task.logfile = logfile
            else:
                task = Task(
                    name="importfromfolder",
                    submitted=now,
                    started=now,
                    status="0%",
                    user=self.user,
                    logfile=logfile,
                )
            task.processid = os.getpid()
            task.save(using=self.database)

            # Choose the right self.delimiter and language
            self.delimiter = (
                get_format("DECIMAL_SEPARATOR", settings.LANGUAGE_CODE, True) == ","
                and ";"
                or ","
            )
            translation.activate(settings.LANGUAGE_CODE)

            # confirm there is enough storage to proceed
            maxstorage = getattr(settings, "MAXSTORAGE", 0) or 0
            if maxstorage:
                storageUsage = round(getStorageUsage() / 1024 / 1024)
                if storageUsage > maxstorage:
                    raise CommandError(
                        "Storage quota exceeded: %sMB used out of %sMB available. Please free some disk space and try again"
                        % (storageUsage, maxstorage)
                    )

            # Execute
            if "FILEUPLOADFOLDER" in get_databases()[self.database] and os.path.isdir(
                get_databases()[self.database]["FILEUPLOADFOLDER"]
            ):
                # Open the logfile
                logger.info("Started importfromfolder\n")

                all_models = [
                    (ct.model_class(), ct.pk)
                    for ct in ContentType.objects.all()
                    if ct.model_class()
                ]
                models = []
                for ifile in os.listdir(
                    get_databases()[self.database]["FILEUPLOADFOLDER"]
                ):
                    if not ifile.lower().endswith(
                        (
                            ".sql",
                            ".sql.gz",
                            ".csv",
                            ".csv.gz",
                            ".cpy",
                            ".cpy.gz",
                            ".xlsx",
                            ".xlsm",
                        )
                    ):
                        continue
                    filename0 = ifile.split(".")[0].split(" (")[0]

                    model = None
                    contenttype_id = None
                    for m, ct in all_models:
                        if matchesModelName(filename0, m):
                            model = m
                            contenttype_id = ct
                            break

                    if not model or model in EXCLUDE_FROM_BULK_OPERATIONS:
                        logger.info("Ignoring data in file: %s" % ifile)
                    elif self.user and not self.user.has_perm(
                        "%s.%s"
                        % (
                            model._meta.app_label,
                            get_permission_codename("add", model._meta),
                        )
                    ):
                        # Check permissions
                        logger.info("You don't have permissions to add: %s" % ifile)
                    else:
                        deps = set([model])
                        GridReport.dependent_models(model, deps)

                        models.append((ifile, model, contenttype_id, deps))

                # Sort the list of models, based on dependencies between models
                models = GridReport.sort_models(models)

                i = 0
                cnt = len(models)
                skip_audit_log = (
                    Parameter.getValue(
                        "import_skips_audit_log", self.database, "true"
                    ).lower()
                    == "true"
                )
                for ifile, model, contenttype_id, dependencies in models:
                    task.status = str(int(i / cnt * 100)) + "%"
                    task.message = "Processing data file %s" % ifile
                    task.save(using=self.database)
                    i += 1
                    filetoparse = os.path.join(
                        os.path.abspath(
                            get_databases()[self.database]["FILEUPLOADFOLDER"]
                        ),
                        ifile,
                    )
                    starting = datetime.now()
                    if ifile.lower().endswith((".sql", ".sql.gz")):
                        logger.info(
                            "Started executing SQL statements from file: %s" % ifile
                        )
                        errors[0] += self.executeSQLfile(filetoparse)
                        logger.info(
                            "Finished executing SQL statements from file %s in %s"
                            % (ifile, timesince(starting))
                        )
                    elif ifile.lower().endswith((".cpy", ".cpy.gz")):
                        logger.info("Started uploading copy file: %s" % ifile)
                        errors[0] += self.executeCOPYfile(model, filetoparse)
                        logger.info(
                            "Finished uploading copy file %s in %s"
                            % (ifile, timesince(starting))
                        )
                    elif ifile.lower().endswith((".xlsx", ".xlsm")):
                        logger.info("Started processing data in Excel file: %s" % ifile)
                        returnederrors = self.loadExcelfile(
                            model, filetoparse, skip_audit_log
                        )
                        errors[0] += returnederrors[0]
                        errors[1] += returnederrors[1]
                        logger.info(
                            "Finished processing data in file %s in %s"
                            % (ifile, timesince(starting))
                        )
                    else:
                        logger.info("Started processing data in CSV file: %s" % ifile)
                        returnederrors = self.loadCSVfile(
                            model, filetoparse, skip_audit_log
                        )
                        errors[0] += returnederrors[0]
                        errors[1] += returnederrors[1]
                        logger.info(
                            "Finished processing data in CSV file %s in %s"
                            % (ifile, timesince(starting))
                        )
            else:
                cnt = 0
                logger.error("Failed, folder does not exist")

            # Task update
            if errors[0] > 0:
                task.status = "Failed"
                task.message = (
                    "Uploaded %s data files with %s errors and %s warnings"
                    % (cnt, errors[0], errors[1])
                )
            else:
                task.status = "Done"
                task.message = "Uploaded %s data files with %s warnings" % (
                    cnt,
                    errors[1],
                )
            task.finished = datetime.now()

        except KeyboardInterrupt:
            if task:
                task.status = "Cancelled"
                task.message = "Cancelled"
            logger.info("Cancelled\n")

        except Exception as e:
            logger.error("Failed")
            if task:
                task.status = "Failed"
                task.message = "%s" % e
            raise e

        finally:
            setattr(_thread_locals, "database", old_thread_locals)
            if task:
                if errors[0] == 0:
                    task.status = "Done"
                else:
                    task.status = "Failed"
                task.processid = None
                task.finished = datetime.now()
                task.save(using=self.database)
            logger.info("End of importfromfolder in %s\n" % timesince(startofall))

    def executeCOPYfile(self, model, ifile):
        """
        Use the copy command to upload data into the database
        The filename must be equal to the table name (E.g : demand, buffer, forecast...)
        The first line of the file must contain the columns to popualate, comma separated
        """
        cursor = connections[self.database].cursor()
        try:
            if ifile.lower().endswith(".gz"):
                file_open = gzip.open
            else:
                file_open = open

            # Retrieve the header line of the file
            tableName = model._meta.db_table
            f = file_open(ifile, "rt")
            firstLine = f.readline().rstrip()
            f.close()

            # Validate the data fields in the header
            headers = []
            for f in firstLine.split(","):
                col = f.strip().strip("#").strip('"').lower() if f else ""
                dbfield = None
                for i in model._meta.fields:
                    # Try with database field name
                    if col == i.get_attname():
                        dbfield = i.get_attname()
                        break
                    # Try with translated field names
                    elif col == i.name.lower() or col == i.verbose_name.lower():
                        dbfield = i.get_attname()
                        break
                    if translation.get_language() != "en":
                        # Try with English field names
                        with translation.override("en"):
                            if col == i.name.lower() or col == i.verbose_name.lower():
                                dbfield = i.get_attname()
                                break
                if dbfield:
                    headers.append('"%s"' % dbfield)
                else:
                    raise Exception("Invalid field name '%s'" % col)

            # count how many records in table before the copy operation
            cursor.execute("select count(*) from %s" % tableName)
            countBefore = cursor.fetchone()[0]

            # Load the data records
            copyFile = file_open(ifile)
            cursor.copy_expert(
                "copy %s (%s) from STDIN with delimiter ',' csv header"
                % (tableName, ",".join(headers)),
                copyFile,
            )

            # count how many records in table after the copy operation
            cursor.execute("select count(*) from %s" % tableName)
            countAfter = cursor.fetchone()[0]

            logger.info(
                "%s records uploaded into table %s"
                % (
                    (countAfter - countBefore),
                    tableName,
                )
            )
            return 0

        except Exception as e:
            logger.error("Error uploading COPY file: %s" % e)
            return 1
        finally:
            # Need to force closing the connection. Otherwise we keep the
            # connection in the restricted role.
            connections[self.database].close()

    def executeSQLfile(self, ifile):
        """
        Execute statements from a text with SQL statements.
        """
        if ifile.lower().endswith(".gz"):
            file_open = gzip.open
        else:
            file_open = open
        try:
            with connections[
                (
                    self.database
                    if f"{self.database}_report" not in get_databases(True)
                    else f"{self.database}_report"
                )
            ].cursor() as cursor:
                cursor.execute(file_open(ifile, "rt").read())

            return 0
        except Exception as e:
            logger.error("Error executing SQL: %s" % e)
            return 1

    def loadCSVfile(self, model, file, skip_audit_log):
        errorcount = 0
        warningcount = 0
        datafile = EncodedCSVReader(file, delimiter=self.delimiter)
        try:
            with transaction.atomic(using=self.database):
                for error in parseCSVdata(
                    model,
                    datafile,
                    user=self.user,
                    database=self.database,
                    skip_audit_log=skip_audit_log,
                ):
                    if error[0] == logging.ERROR:
                        logger.error(
                            "Error: %s%s%s%s"
                            % (
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )
                        errorcount += 1
                    elif error[0] == logging.WARNING:
                        logger.warning(
                            "Warning: %s%s%s%s"
                            % (
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )
                        warningcount += 1
                    else:
                        logger.info(
                            "%s%s%s%s"
                            % (
                                "Row %s: " % error[1] if error[1] else "",
                                "field %s: " % error[2] if error[2] else "",
                                "%s: " % error[3] if error[3] else "",
                                error[4],
                            )
                        )

            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(database=self.database, url=None)

        except Exception:
            errorcount += 1
            logger.error("Error: Invalid data format - skipping the file \n")
        return [errorcount, warningcount]

    def loadExcelfile(self, model, file, skip_audit_log):
        errorcount = 0
        warningcount = 0

        # retrieve value of parameter excel_duration_in_days
        excel_duration_in_days = (
            Parameter.getValue("excel_duration_in_days", self.database, "false").lower()
            == "true"
        )

        try:
            with transaction.atomic(using=self.database):
                wb = load_workbook(filename=file, data_only=True)
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    for error in parseExcelWorksheet(
                        model,
                        ws,
                        user=self.user,
                        database=self.database,
                        excel_duration_in_days=excel_duration_in_days,
                        skip_audit_log=skip_audit_log,
                    ):
                        if error[0] == logging.ERROR:
                            logger.error(
                                "Error: %s%s%s%s"
                                % (
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
                            errorcount += 1
                        elif error[0] == logging.WARNING:
                            logger.warning(
                                "Warning: %s%s%s%s"
                                % (
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
                            warningcount += 1
                        else:
                            logger.info(
                                "%s%s%s%s"
                                % (
                                    "Row %s: " % error[1] if error[1] else "",
                                    "field %s: " % error[2] if error[2] else "",
                                    "%s: " % error[3] if error[3] else "",
                                    error[4],
                                )
                            )
            # Records are committed. Launch notification generator now.
            NotificationFactory.launchWorker(database=self.database, url=None)
        except Exception:
            errorcount += 1
            logger.error("Error: Invalid data format - skipping the file \n")
        return [errorcount, warningcount]

    def email_info(self, database):
        scenario = "" if database == DEFAULT_DB_ALIAS else f"/{database}"
        return f'<a href="{settings.EMAIL_URL_PREFIX}{scenario}/execute/downloadfromfolder/0/">Download import Files</a>'

    # accordion template
    title = _("Import data files")
    index = 1100
    help_url = "command-reference.html#importfromfolder"

    @staticmethod
    def getHTML(request):
        if (
            "FILEUPLOADFOLDER" not in get_databases()[request.database]
            or not request.user.is_superuser
        ):
            return None

        filestoupload = []
        if "FILEUPLOADFOLDER" in get_databases()[request.database]:
            uploadfolder = get_databases()[request.database]["FILEUPLOADFOLDER"]
            if os.path.isdir(uploadfolder):
                tzoffset = GridReport.getTimezoneOffset(request)
                for file in sorted(os.listdir(uploadfolder)):
                    if file.endswith(
                        (
                            ".csv",
                            ".csv.gz",
                            ".xlsx",
                            ".xlsm",
                            ".cpy",
                            ".sql",
                            ".cpy.gz",
                            ".sql.gz",
                        )
                    ):
                        stat = os.stat(os.path.join(uploadfolder, file))
                        filestoupload.append(
                            [
                                file,
                                strftime(
                                    "%Y-%m-%d %H:%M:%S",
                                    localtime(stat.st_mtime + tzoffset.total_seconds()),
                                ),
                                sizeof_fmt(stat.st_size),
                            ]
                        )

        return render_to_string(
            "commands/importfromfolder.html",
            {"filestoupload": filestoupload},
            request=request,
        )


class EncodedCSVReader:
    """
    A CSV reader which will iterate over lines in the CSV data buffer.
    The reader will scan the BOM header in the data to detect the right encoding.
    """

    def __init__(self, datafile, **kwds):
        # Read the file into memory
        # TODO Huge file uploads can overwhelm your system!

        # Detect the encoding of the data by scanning the BOM.
        # Skip the BOM header if it is found.

        if datafile.lower().endswith(".gz"):
            file_open = gzip.open
        else:
            file_open = open
        self.reader = file_open(datafile, "rb")
        data = self.reader.read(5)
        self.reader.close()
        if data.startswith(codecs.BOM_UTF32_BE):
            self.reader = file_open(datafile, "rt", encoding="utf_32_be")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF32_LE):
            self.reader = file_open(datafile, "rt", encoding="utf_32_le")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_BE):
            self.reader = file_open(datafile, "rt", encoding="utf_16_be")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF16_LE):
            self.reader = file_open(datafile, "rt", encoding="utf_16_le")
            self.reader.read(1)
        elif data.startswith(codecs.BOM_UTF8):
            self.reader = file_open(datafile, "rt", encoding="utf_8")
            self.reader.read(1)
        else:
            # No BOM header found. We assume the data is encoded in the default CSV character set.
            self.reader = file_open(datafile, "rt", encoding=settings.CSV_CHARSET)

        # Open the file
        self.csvreader = csv.reader(self.reader, **kwds)

    def __next__(self):
        return next(self.csvreader)

    def __iter__(self):
        return self
