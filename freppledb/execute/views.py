#
# Copyright (C) 2007-2020 by frePPLe bv
#
# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE
# LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION
# OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION
# WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#

from datetime import datetime, timedelta
from importlib import import_module
from io import BytesIO
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.comments import Comment as CellComment
import os
import psutil
import re
import shlex
from time import sleep, localtime, strftime
from zipfile import ZipFile, ZIP_DEFLATED

from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_permission_codename
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.db.models.fields import AutoField
from django.db.models.fields.related import ForeignKey
from django.views.decorators.cache import never_cache
from django.shortcuts import render
from django.utils.html import escape
from django.utils.translation import gettext_lazy as _
from django.db import DEFAULT_DB_ALIAS
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.http import (
    Http404,
    HttpResponseRedirect,
    HttpResponseServerError,
    HttpResponse,
    HttpResponseNotFound,
    HttpResponseForbidden,
    HttpResponseNotAllowed,
    JsonResponse,
    StreamingHttpResponse,
)
from django.contrib import messages
from django.utils.encoding import force_str
from django.utils.text import capfirst
from django.core.management import get_commands, call_command

from freppledb import __version__
from freppledb.admin import data_site
from freppledb.common.dataload import parseExcelWorksheet
from freppledb.common.models import Scenario, HierarchyModel, Parameter
from freppledb.common.report import (
    GridFieldDuration,
    GridFieldBool,
    GridFieldLocalDateTime,
    GridReport,
    GridFieldText,
    GridFieldInteger,
    EXCLUDE_FROM_BULK_OPERATIONS,
    _getCellValue,
    matchesModelName,
    sizeof_fmt,
)
from freppledb.common.utils import get_databases

from freppledb.common.utils import forceWsgiReload
from freppledb.common.views import sendStaticFile
from .utils import updateScenarioCount
from .models import Task, ScheduledTask, DataExport
from .management.commands.runworker import launchWorker
from .management.commands.runplan import parseConstraints, constraintString
from .management.commands.scheduletasks import scheduler

import logging


logger = logging.getLogger(__name__)


class TaskReport(GridReport):
    """
    A list report to review the history of actions.
    """

    template = "execute/execute.html"
    title = _("Task status")
    basequeryset = (
        Task.objects.all()
        .extra(
            select={
                "duration": """case when processid is not null or started is not null then
                   date_trunc('second', coalesce(finished::timestamp(0), now()) - started::timestamp(0))
                   end
                   """
            }
        )
        .select_related("user")
    )
    model = Task
    frozenColumns = 0
    multiselect = False
    editable = False
    height = 150
    default_sort = (0, "desc")
    help_url = "user-interface/execute.html"

    rows = (
        GridFieldInteger("id", title=_("identifier"), key=True),
        GridFieldText("name", title=_("name"), editable=False, align="center"),
        GridFieldLocalDateTime(
            "submitted", title=_("submitted"), editable=False, align="center"
        ),
        GridFieldLocalDateTime(
            "started", title=_("started"), editable=False, align="center"
        ),
        GridFieldLocalDateTime(
            "finished", title=_("finished"), editable=False, align="center"
        ),
        GridFieldDuration(
            "duration",
            title=_("duration"),
            search=False,
            editable=False,
            align="center",
        ),
        GridFieldText(
            "status",
            title=_("status"),
            editable=False,
            align="center",
            extra="formatter:status",
        ),
        GridFieldText(
            "logfile",
            title=_("log file"),
            width=80,
            editable=False,
            align="center",
            extra="formatter:logbutton",
        ),
        GridFieldText(
            "message",
            title=_("message"),
            editable=False,
            width=500,
            formatter="longstring",
        ),
        GridFieldText(
            "arguments", title=_("arguments"), formatter="longstring", editable=False
        ),
        GridFieldText(
            "user",
            title=_("user"),
            field_name="user__username",
            editable=False,
            align="center",
        ),
        GridFieldBool("cancelable", title="cancelable", hidden=True),
    )

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        # Get the list of all available commands
        all_commands = []
        cnt = 0
        for commandname, appname in get_commands().items():
            try:
                cmd = getattr(
                    import_module("%s.management.commands.%s" % (appname, commandname)),
                    "Command",
                )
                if getattr(cmd, "index", -1) >= 0 and getattr(cmd, "getHTML", None):
                    cmd.name = commandname
                    html = cmd.getHTML(request)
                    if html:
                        all_commands.append(
                            {
                                "command": cmd,
                                "options": {"collapsed": cmd.name != "runplan"},
                                "html": html,
                            }
                        )
                    cnt += 1
            except Exception as e:
                logger.warning(
                    "Couldn't import getHTML method from %s.management.commands.%s: %s"
                    % (appname, commandname, e)
                )

        # Use the preferences
        commandlist1 = []
        commandlist2 = []
        prefs = getattr(request, "prefs", None)
        if prefs:
            # Add preferences to the command lists
            widgets = prefs.get("widgets", [])
            for row in widgets:
                column = row.get("name", "")
                if column == "column1":
                    column = commandlist1
                elif column == "column2":
                    column = commandlist2
                else:
                    continue
                for c in row.get("cols", []):
                    for w in c.get("widgets", []):
                        cnt1 = 0
                        for cmd in all_commands:
                            if cmd["command"].name == w[0]:
                                cmd["options"] = w[1]
                                cnt += 1
                                column.append(cmd)
                                del all_commands[cnt1]
                                break
                            cnt1 += 1
            # Add any commands not yet in the preferences
            if widgets:
                for w in all_commands:
                    commandlist1.append(w)
        if not commandlist1 and not commandlist2:
            # No preferences, divide the commands equal over 2 columns
            all_commands.sort(key=lambda x: x["command"].index)
            mid = (len(all_commands) + 1) // 2
            commandlist1 = all_commands[:mid]
            commandlist2 = all_commands[mid:]

        # Send to template
        return {"commandlist1": commandlist1, "commandlist2": commandlist2}

    @classmethod
    def query(reportclass, request, basequery, sortsql="1 asc"):
        logfileslist = set(
            [
                x
                for x in os.listdir(settings.FREPPLE_LOGDIR)
                if x.endswith(".log")
                or (x.lower().endswith(".dump") and request.user.is_superuser)
            ]
        )

        # Cancel waiting tasks if no runworker is active
        worker_alive = Parameter.getValue("Worker alive", request.database, None)
        try:
            if not worker_alive or datetime.now() - datetime.strptime(
                worker_alive, "%Y-%m-%d %H:%M:%S"
            ) > timedelta(0, 30):
                Task.objects.using(request.database).filter(
                    status__iexact="waiting",
                    submitted__lte=datetime.now() - timedelta(0, 30),
                ).update(status="Canceled")
        except Exception:
            pass

        for rec in basequery:
            yield {
                "id": rec.id,
                "name": rec.name,
                "submitted": rec.submitted,
                "started": rec.started,
                "finished": rec.finished,
                "status": rec.status,
                "logfile": rec.logfile if rec.logfile in logfileslist else None,
                "message": rec.message,
                "arguments": rec.arguments,
                "user__username": rec.user.username if rec.user else None,
                "duration": rec.duration,
                "cancelable": rec.processid is not None or rec.status == "Waiting",
            }

    @classmethod
    def extraJSON(reportclass, request):
        try:
            lastCompletedTask = (
                Task.objects.all()
                .using(request.database)
                .filter(status="Done")
                .order_by("-id")
                .only("id")[0]
            )
            return '"lastcompleted":%d,\n' % lastCompletedTask.id
        except Exception:
            return '"lastcompleted":0,\n'


@csrf_exempt
def APITask(request, action):
    try:
        if action == "status":
            response = {}
            args = request.GET.getlist("id")
            if args:
                msg = "retrieving task status of %s" % ",".join(args)
                tasks = Task.objects.all().using(request.database).filter(id__in=args)
            else:
                msg = "retrieving task status of all running tasks"
                tasks = (
                    Task.objects.all()
                    .using(request.database)
                    .filter(finished__isnull=True)
                    .exclude(status="Canceled")
                )
            for t in tasks:
                response[t.id] = {
                    "name": t.name,
                    "submitted": str(t.submitted),
                    "started": str(t.started),
                    "finished": str(t.finished),
                    "arguments": t.arguments,
                    "status": t.status,
                    "message": t.message,
                    "user": t.user.username if t.user else None,
                    "logfile": t.logfile,
                }
        elif action == "cancel":
            response = {}
            with transaction.atomic(using=request.database):
                args = request.GET.getlist("id")
                msg = "canceling task %s" % ",".join(args)
                for t in (
                    Task.objects.all()
                    .using(request.database)
                    .select_for_update()
                    .filter(id__in=args)
                ):
                    if request.user.is_superuser or t.user == request.user:
                        if t.processid:
                            # Kill the process with signal 9
                            child_pid = [
                                c.pid for c in psutil.Process(t.processid).children()
                            ]
                            os.kill(t.processid, 9)
                            for child_task in (
                                Task.objects.all()
                                .using(request.database)
                                .filter(processid__in=child_pid)
                            ):
                                try:
                                    os.kill(child_task.processid, 9)
                                except Exception:
                                    pass
                                child_task.message = "Canceled process"
                                child_task.processid = None
                                child_task.status = "Canceled"
                                child_task.save(using=request.database)
                            sleep(1)  # Wait for it to die
                            t.message = "Canceled process"
                            t.processid = None
                        elif t.status != "Waiting":
                            continue
                        t.status = "Canceled"
                        t.save(using=request.database)
                        response[t.id] = {
                            "name": t.name,
                            "submitted": str(t.submitted),
                            "started": str(t.started),
                            "finished": str(t.finished),
                            "arguments": t.arguments,
                            "status": t.status,
                            "message": t.message,
                            "user": t.user.username if t.user else None,
                        }
        elif action == "log":
            taskid = request.GET.get("id")
            if not taskid:
                return HttpResponseNotFound("Task or log file not found")
            try:
                msg = "retrieving task log of %s" % taskid
                task = Task.objects.all().using(request.database).get(id=taskid)
                if task and task.logfile:
                    return sendStaticFile(
                        request,
                        settings.FREPPLE_LOGDIR,
                        task.logfile,
                        headers={
                            "Content-Type": "application/octet-stream",
                            "Content-Disposition": 'inline; filename="frepple_%s_%s.log"'
                            % (request.database, taskid),
                        },
                    )
            except Exception as e:
                return HttpResponseNotFound("Task or log file not found")
        else:
            msg = "launching task"
            task = wrapTask(request, action)
            if task:
                response = {"taskid": task.id, "message": "Successfully launched task"}
            else:
                response = {"message": "No task was launched"}
    except Exception:
        response = {"message": "Exception %s" % msg}
    return HttpResponse(json.dumps(response), content_type="application/json")


@staff_member_required
@never_cache
@csrf_protect
def LaunchTask(request, action):
    try:
        if action == "exportworkbook":
            return exportWorkbook(request)
        elif action == "importworkbook":
            return StreamingHttpResponse(
                content_type="text/plain; charset=%s" % settings.DEFAULT_CHARSET,
                streaming_content=importWorkbook(request),
            )
        elif action in ("frepple_stop_web_service", "stopwebservice"):
            if not request.user.has_perm("auth.auth.generate_plan"):
                raise Exception("Missing execution privileges")
            from django.core import management

            management.call_command(
                "stopwebservice", force=True, database=request.database
            )
            return HttpResponseRedirect("%s/execute/" % request.prefix)
        else:
            wrapTask(request, action)
            return HttpResponseRedirect("%s/execute/" % request.prefix)
    except Exception as e:
        logger.error("Error launching task: %s" % e)
        messages.add_message(
            request, messages.ERROR, force_str(_("Failure launching task"))
        )
        return HttpResponseRedirect("%s/execute/" % request.prefix)


def wrapTask(request, action):
    # Allow only post
    if request.method != "POST":
        raise Exception("Only post requests allowed")
    # Parse the posted parameters as arguments for an asynchronous task to add to the queue.    TODO MAKE MODULAR WITH SEPERATE TASK CLASS
    worker_database = request.database

    now = datetime.now()
    task = None
    args = request.POST or request.GET

    # A
    # TODO remove special case - call runwebservice or runplan instead
    if action in ("runplan", "runwebservice"):
        if not request.user.has_perm("auth.auth.generate_plan"):
            raise Exception("Missing execution privileges")
        constraint = parseConstraints(",".join(args.getlist("constraint")))
        task = Task(name="runplan", submitted=now, status="Waiting", user=request.user)
        background = False
        if action in ("frepple_start_web_service", "runwebservice"):
            # Load existing plan and run as a web service
            background = True
            env = []
            constraint = 4 + 16 + 32
            plantype = 1
            try:
                lastrun = (
                    Task.objects.all()
                    .using(request.database)
                    .filter(name="runplan")
                    .order_by("-id")[0]
                )
                for i in shlex.split(lastrun.arguments or ""):
                    if "=" in i:
                        key, val = i.split("=")
                        if key == "--constraint":
                            constraint = parseConstraints(val)
                        elif key == "--plantype":
                            plantype = int(val)
            except Exception:
                pass
            task.arguments = "--constraint=%s --plantype=%s --background" % (
                constraintString(constraint),
                plantype,
            )
        else:
            # Create a new plan
            task.arguments = "--constraint=%s --plantype=%s" % (
                constraintString(constraint),
                args.get("plantype", 1),
            )
            env = []
            for value in args.getlist("env"):
                env.append(value)
            task.arguments = "--constraint=%s --plantype=%s" % (
                constraintString(constraint),
                args.get("plantype", 1),
            )
            if (
                Parameter.getValue("plan.webservice", request.database, "true").lower()
                == "true"
            ):
                task.arguments += " --background"
                background = True
        if background:
            # Avoid getting multiple waiting tasks in background mode
            cnt = Task.objects.all().filter(name="runplan", status="Waiting").count()
            if cnt > 0:
                return None
        if env:
            task.arguments += " --env=%s" % ",".join(env)
        else:
            task.arguments += " --env=loadplan"
        task.save(using=request.database)
    # C
    elif action == "empty":
        if not request.user.has_perm("auth.run_db"):
            raise Exception("Missing execution privileges")
        task = Task(name="empty", submitted=now, status="Waiting", user=request.user)
        models = ",".join(args.getlist("models"))
        if models:
            task.arguments = "--models=%s" % (models)
        elif "all" in args:
            task.arguments = "--all"
        task.save(using=request.database)
    # D
    elif action == "loaddata":
        if not request.user.has_perm("auth.run_db"):
            raise Exception("Missing execution privileges")

        # Step 1: optionally empty the database
        if args.get("emptybefore", False) == "true":
            task = Task(
                name="empty",
                submitted=now,
                status="Waiting",
                user=request.user,
                arguments="--all",
            )
            task.save(using=request.database)

        # Step 2: load the fixture
        task = Task(
            name="loaddata",
            submitted=now,
            status="Waiting",
            user=request.user,
            arguments=args["fixture"],
        )
        task.save(using=request.database)

        # Step 3: optionally run the plan
        if args.get("regenerateplan", False) == "true":
            active_modules = "supply"
            if "freppledb.forecast" in settings.INSTALLED_APPS:
                active_modules = "fcst," + active_modules
            task = Task(
                name="runplan", submitted=now, status="Waiting", user=request.user
            )
            task.arguments = (
                "--constraint=capa,mfg_lt,po_lt --plantype=1 --env=%s --background"
                % (active_modules,)
            )
            task.save(using=request.database)
    # E
    elif action == "scenario_copy":
        worker_database = DEFAULT_DB_ALIAS
        if "copy" in args:
            if not request.user.has_perm("common.copy_scenario"):
                raise Exception("Missing execution privileges")
            source = args.get("source", request.database)
            worker_database = source
            destination = args.get("destination", False)
            dumpfile = args.get("dumpfile")
            if destination and destination != DEFAULT_DB_ALIAS:
                force = args.get("force", False)
                arguments = "%s %s" % (source, destination)
                if dumpfile:
                    arguments = "--dumpfile=%s %s" % (dumpfile, arguments)
                if force:
                    arguments += " --force"
                task = Task(
                    name="scenario_copy",
                    submitted=now,
                    status="Waiting",
                    user=request.user,
                    arguments=arguments,
                )
                task.save(using=source)
        elif "release" in args:
            if not request.user.has_perm("common.release_scenario"):
                raise Exception("Missing execution privileges")

            worker_database = request.database

            if request.database != DEFAULT_DB_ALIAS:
                task = Task(
                    name="scenario_release",
                    submitted=now,
                    status="Waiting",
                    user=request.user,
                    arguments="--database=%s"
                    % (args["database"] if "database" in args else request.database,),
                )
                task.save(using=request.database)
        elif "promote" in args:
            if not request.user.has_perm("common.promote_scenario"):
                raise Exception("Missing execution privileges")
            source = args.get("source", request.database)
            worker_database = source
            destination = args.get("destination", False)
            if destination and destination == DEFAULT_DB_ALIAS:
                arguments = "--promote %s %s" % (source, destination)
                task = Task(
                    name="scenario_copy",
                    submitted=now,
                    status="Waiting",
                    user=request.user,
                    arguments=arguments,
                )
                task.save(using=source)
        elif "update" in args:
            # Note: update is immediate and synchronous.
            if not request.user.has_perm("common.release_scenario"):
                raise Exception("Missing execution privileges")
            scenario = args["scenario"]
            sc = Scenario.objects.using(DEFAULT_DB_ALIAS).get(name=scenario)
            sc.description = args.get("description", None)
            sc.save(update_fields=["description"], using=DEFAULT_DB_ALIAS)
            return HttpResponse(content="OK")
        elif "theme" in args:
            # Note: update is immediate and synchronous.
            scenario = args.get("scenario", "").strip()
            theme = args["theme"].strip()
            if theme not in settings.THEMES:
                theme = None
            sc = Scenario.objects.using(DEFAULT_DB_ALIAS).get(name=scenario)
            if theme:
                if request.user.scenario_themes:
                    request.user.scenario_themes[scenario] = theme
                else:
                    request.user.scenario_themes = {scenario: theme}
            else:
                del request.user.scenario_themes[scenario]
            request.user.save(update_fields=["scenario_themes"], using=DEFAULT_DB_ALIAS)
            return HttpResponse(content="OK")
        else:
            raise Exception("Invalid scenario task")
    # G
    elif action == "createbuckets":
        if not request.user.has_perm("auth.run_db"):
            raise Exception("Missing execution privileges")
        task = Task(
            name="createbuckets", submitted=now, status="Waiting", user=request.user
        )
        arguments = []
        weekstart = args.get("weekstart", None)
        if weekstart:
            arguments.append("--weekstart=%s" % weekstart)
        format_day = args.get("format-day", None)
        if format_day:
            arguments.append('--format-day="%s"' % format_day)
        format_week = args.get("format-week", None)
        if format_week:
            arguments.append('--format-week="%s"' % format_week)
        format_month = args.get("format-month", None)
        if format_month:
            arguments.append('--format-month="%s"' % format_month)
        format_quarter = args.get("format-quarter", None)
        if format_quarter:
            arguments.append('--format-quarter="%s"' % format_quarter)
        format_year = args.get("format-year", None)
        if format_year:
            arguments.append('--format-year="%s"' % format_year)
        if arguments:
            task.arguments = " ".join(arguments)
        task.save(using=request.database)
    elif action == "measure_copy":
        if not request.user.has_perm("auth.run_db"):
            raise Exception("Missing execution privileges")
        task = Task(
            name="measure_copy",
            submitted=now,
            status="Waiting",
            user=request.user,
            arguments="%s %s" % (args["source"], args["destination"]),
        )
        task.save(using=request.database)
    else:
        # Generic task wrapper

        # Find the command and verify we have permissions to run it
        command = None
        for commandname, appname in get_commands().items():
            if commandname == action:
                try:
                    c = getattr(
                        import_module(
                            "%s.management.commands.%s" % (appname, commandname)
                        ),
                        "Command",
                    )
                    if c.index >= 0:
                        if getattr(c, "getHTML", None) and c.getHTML(request):
                            # Command class has getHTML method
                            command = c
                            break
                        else:
                            for p in c.__bases__:
                                # Parent command class has getHTML method
                                if getattr(p, "getHTML", None) and p.getHTML(request):
                                    command = c
                                    break
                            if command:
                                break
                except Exception:
                    pass  # Silently ignore failures
        if not command:
            raise Exception("Invalid task name '%s'" % action)
        # Create a task
        arguments = []
        for arg, val in args.lists():
            if arg != "csrfmiddlewaretoken":
                arguments.append("--%s=%s" % (arg, shlex.quote(",".join(val))))
        task = Task(name=action, submitted=now, status="Waiting", user=request.user)
        if arguments:
            task.arguments = " ".join(arguments)
        task.save(using=request.database)

    # Launch a worker process
    if task:
        launchWorker(database=worker_database)
    return task


@staff_member_required
@never_cache
@csrf_protect
def CancelTask(request, taskid):
    # Allow only post
    if (
        request.method != "POST"
        or request.headers.get("x-requested-with") != "XMLHttpRequest"
    ):
        raise Http404("Only ajax post requests allowed")
    try:
        task = Task.objects.all().using(request.database).get(pk=taskid)
        if task.processid:
            # Kill the process with signal 9
            child_pid = [c.pid for c in psutil.Process(task.processid).children()]
            os.kill(task.processid, 9)
            for child_task in (
                Task.objects.all()
                .using(request.database)
                .filter(processid__in=child_pid)
            ):
                try:
                    os.kill(child_task.processid, 9)
                except Exception:
                    pass
                child_task.message = "Canceled process"
                child_task.processid = None
                child_task.status = "Canceled"
                child_task.save(using=request.database)
            task.message = "Canceled process"
            task.processid = None
        elif task.status != "Waiting":
            return HttpResponseServerError("Task isn't running or waiting to run")
        task.status = "Canceled"
        task.save(using=request.database)
        # Just in case, to cover corner cases
        launchWorker(database=request.database)
        return HttpResponse(content="OK")
    except PermissionError:
        return HttpResponseServerError("No permission to kill this task")
    except Exception as e:
        # We don't know why it failed. We just clean things up.
        task.message = "Canceled process"
        task.processid = None
        task.status = "Canceled"
        task.save(using=request.database)
        return HttpResponse(content="OK")


@staff_member_required
@never_cache
def DownloadLogFile(request, taskid):
    filename = Task.objects.using(request.database).get(id=taskid).logfile
    if (
        filename.lower().endswith(".dump") and not request.user.is_superuser
    ) or not filename.lower().endswith((".log", ".dump")):
        return HttpResponseNotFound(force_str(_("Error")))
    return sendStaticFile(
        request,
        settings.FREPPLE_LOGDIR,
        filename,
        headers={
            "Content-Type": "application/octet-stream",
            "Content-Disposition": 'inline; filename="%s"' % filename,
        },
    )


@staff_member_required
@never_cache
def DeleteLogFile(request, taskid):
    if request.method != "POST":
        return HttpResponseNotAllowed(
            ["post"], content="Only POST request method is allowed"
        )
    filename = Task.objects.using(request.database).get(id=taskid).logfile
    if (
        not filename
        or (filename.lower().endswith(".dump") and not request.user.is_superuser)
        or not filename.lower().endswith((".log", ".dump"))
    ):
        return HttpResponseNotFound(force_str(_("Error")))
    try:
        os.remove(os.path.join(settings.FREPPLE_LOGDIR, filename))
        Task.objects.using(request.database).filter(id=taskid).update(logfile=None)
        return HttpResponse(content="OK")
    except Exception as e:
        logger.error("Couldn't delete log file: %s" % e)
        return HttpResponseForbidden(force_str(_("Couldn't delete file")))


@staff_member_required
@never_cache
def logfile(request, taskid):
    """
    This view shows the frePPLe log file of the last planning run in this database.
    """
    filesize = None
    try:
        filename = Task.objects.using(request.database).get(id=taskid).logfile
        if not filename.lower().endswith(".log"):
            return HttpResponseNotFound(force_str(_("Error")))

        f = open(os.path.join(settings.FREPPLE_LOGDIR, filename), "rb")
    except Exception:
        logdata = "File not found"
    else:
        try:
            f.seek(-1, os.SEEK_END)
            filesize = f.tell()
            if filesize >= 50000:
                # Too big to display completely
                f.seek(-50000, os.SEEK_END)
                d = f.read(50000)
                d = d[d.index(b"\n") :]  # Strip the first, incomplete line
                logdata = (
                    force_str(_("Displaying only the last 50K from the log file"))
                    + "...\n\n..."
                    + d.decode("utf8", "ignore")
                )
            else:
                # Displayed completely
                f.seek(0, os.SEEK_SET)
                logdata = f.read(50000).decode("utf8", "ignore")
        finally:
            f.close()

    return render(
        request,
        "execute/logfrepple.html",
        {
            "title": " ".join([force_str(capfirst(_("log file"))), taskid]),
            "logdata": logdata,
            "taskid": taskid,
            "filesize": sizeof_fmt(filesize),
        },
    )


class FileManager:
    """
    Class to upload and download files from a folder.
    The folder code argument indicates which folder to use:
    - 0: file upload folder
    - 1: export subdirectory of the file upload folder
    """

    all_extensions = (
        ".xlsx",
        ".xlsm",
        ".csv",
        ".csv.gz",
        ".cpy",
        ".cpy.gz",
        ".sql",
        ".sql.gz",
    )

    @staticmethod
    def getFolderInfo(request, foldercode):
        if foldercode == "0":
            # File upload folder
            return (
                get_databases()[request.database]["FILEUPLOADFOLDER"],
                FileManager.all_extensions,
            )
        elif foldercode == "1":
            # Export folder
            return (
                os.path.join(
                    get_databases()[request.database]["FILEUPLOADFOLDER"], "export"
                ),
                None,  # No upload here
            )
        else:
            raise Http404("Invalid folder code")

    @staticmethod
    def cleanFolder(foldercode, database=DEFAULT_DB_ALIAS):
        if foldercode == 0:
            folder = get_databases()[database]["FILEUPLOADFOLDER"]
            extensions = FileManager.all_extensions
        elif foldercode == 1:
            folder = os.path.join(
                get_databases()[database]["FILEUPLOADFOLDER"], "export"
            )
            extensions = None
        else:
            raise Exception("Invalid folder code")
        if os.path.isdir(folder):
            for filename in os.listdir(folder):
                if (
                    not extensions or filename.lower().endswith(extensions)
                ) and not filename.lower().endswith(".log"):
                    try:
                        os.remove(os.path.join(folder, filename))
                    except Exception:
                        pass

    @staticmethod
    @csrf_exempt
    @staff_member_required
    @never_cache
    def uploadFiletoFolder(request, foldercode):

        if request.method == "GET":
            # Return a list of available data files
            folder, extensions = FileManager.getFolderInfo(request, foldercode)
            filelist = {}
            if os.path.isdir(get_databases()[request.database]["FILEUPLOADFOLDER"]):
                for filename in os.listdir(folder):
                    clean_filename = re.split(r"/|:|\\", filename)[-1]
                    if (
                        not extensions or clean_filename.lower().endswith(extensions)
                    ) and not clean_filename.lower().endswith(".log"):
                        stat = os.stat(os.path.join(folder, clean_filename))
                        filelist[clean_filename] = {
                            "size_bytes": stat.st_size,
                            "timestamp_utc": strftime(
                                "%Y-%m-%d %H:%M:%S", localtime(stat.st_mtime)
                            ),
                        }

            return JsonResponse(filelist)
        elif request.method in ("POST", "PUT"):
            # Upload a new data file
            if len(list(request.FILES.items())) == 0:
                return HttpResponseNotFound("Missing file selection in request")
            errorcount = 0
            response = HttpResponse()
            folder, extensions = FileManager.getFolderInfo(request, foldercode)

            # Try to create the upload if doesn't exist yet
            if not os.path.isdir(get_databases()[request.database]["FILEUPLOADFOLDER"]):
                try:
                    os.makedirs(get_databases()[request.database]["FILEUPLOADFOLDER"])
                except Exception:
                    errorcount += 1
                    response.write("Upload folder doesn't exist")

            if not errorcount:
                # Directory exists and we can upload files into it
                for filename, content in request.FILES.items():
                    try:
                        # Validate file name
                        clean_filename = re.split(r"/|:|\\", filename)[-1]
                        cleanpath = os.path.normpath(
                            os.path.join(folder, clean_filename)
                        )
                        if (
                            not cleanpath.startswith(folder)
                            or not extensions
                            or not clean_filename.lower().endswith(extensions)
                        ):
                            logger.error(
                                "Failed file upload: incorrect file name '%s'"
                                % filename
                            )
                            response.write(
                                "%s: <strong>Error</strong> %s<br>\n"
                                % (
                                    escape(clean_filename),
                                    _("Filename extension must be among %(ext)s")
                                    % {"ext": ", ".join(extensions)},
                                )
                            )
                            errorcount += 1
                            continue

                        # Write to a file
                        with open(cleanpath, "wb") as thetarget:
                            for chunk in content.chunks():
                                thetarget.write(chunk)

                        response.write(
                            force_str(
                                "%s: <strong>%s</strong><br>"
                                % (escape(clean_filename), _("OK"))
                            )
                        )
                    except Exception as e:
                        logger.error("Failed file upload: %s" % e)
                        response.write(
                            "%s: <strong>Error</strong> %s<br>"
                            % (escape(clean_filename), _("Upload failed"))
                        )
                        errorcount += 1
            if errorcount:
                response.status_code = 400
                response.reason_phrase = (
                    "%s files failed to upload correctly" % errorcount
                )
            return response
        else:
            return HttpResponseNotAllowed(
                ["post", "put", "get"],
                content="Only GET, PUT and POST request methods are allowed",
            )

    @staticmethod
    @csrf_exempt
    @staff_member_required
    @never_cache
    def deleteFilefromFolder(request, foldercode, files):
        if request.method != "DELETE":
            return HttpResponseNotAllowed(
                ["delete"], content="Only DELETE request method is allowed"
            )
        folder, extensions = FileManager.getFolderInfo(request, foldercode)
        if extensions is None:
            extensions = FileManager.all_extensions

        fileerrors = force_str(_("Error deleting file"))
        errorcount = 0
        filelist = list()
        if files == "AllFiles":
            if os.path.isdir(folder):
                for filename in os.listdir(folder):
                    clean_filename = re.split(r"/|:|\\", filename)[-1]
                    if clean_filename.lower().endswith(
                        extensions
                    ) and not clean_filename.lower().endswith(".log"):
                        filelist.append(clean_filename)
        else:
            clean_filename = re.split(r"/|:|\\", files)[-1]
            if os.path.isdir(folder):
                filelist.append(clean_filename)

        for clean_filename in filelist:
            try:
                cleanpath = os.path.normpath(os.path.join(folder, clean_filename))
                if cleanpath.startswith(folder):
                    os.remove(cleanpath)
            except FileNotFoundError:
                # No error message needed
                pass
            except Exception as e:
                logger.error("Failed file deletion: %s" % e)
                errorcount += 1
                fileerrors = fileerrors + " / " + escape(clean_filename)

        if errorcount > 0:
            return HttpResponseServerError(fileerrors)
        else:
            return HttpResponse(content="OK")

    @staticmethod
    @csrf_exempt
    @staff_member_required
    @never_cache
    def downloadFilefromFolder(request, foldercode, filename=None):
        if request.method != "GET":
            return HttpResponseNotAllowed(
                ["get"], content="Only GET request method is allowed"
            )
        folder, extensions = FileManager.getFolderInfo(request, foldercode)
        if not extensions:
            extensions = FileManager.all_extensions
        if filename:
            # Download a single file
            try:
                clean_filename = re.split(r"/|:|\\", filename)[0]
                cleanpath = os.path.normpath(os.path.join(folder, clean_filename))
                if not cleanpath.startswith(folder):
                    logger.warning("Failed file download: %s" % filename)
                    return HttpResponseNotFound(force_str(_("Error")))
                if not os.path.isfile(cleanpath):
                    logger.warning("Failed file download: %s" % filename)
                    return HttpResponseNotFound(force_str(_("Error")))
                return sendStaticFile(
                    request,
                    folder,
                    clean_filename,
                    headers={
                        "Content-Type": "application/octet-stream",
                        "Content-Disposition": 'inline; filename="%s"' % clean_filename,
                    },
                )
            except Exception as e:
                logger.error("Failed file download: %s" % e)
                return HttpResponseNotFound(force_str(_("Error")))
        else:
            # Download all files
            b = BytesIO()
            with ZipFile(file=b, mode="w", compression=ZIP_DEFLATED) as zf:
                if os.path.isdir(folder):
                    for filename in os.listdir(folder):
                        fullfilename = os.path.join(folder, filename)
                        if filename.endswith(extensions) and os.access(
                            fullfilename, os.R_OK
                        ):
                            zf.write(
                                filename=fullfilename,
                                arcname=os.path.basename(filename),
                            )
            response = HttpResponse(b.getvalue(), content_type="application/zip")
            response["Content-Disposition"] = 'attachment; filename="frepple.zip"'
            return response


@staff_member_required
@never_cache
def scheduletasks(request):
    if request.headers.get(
        "x-requested-with"
    ) != "XMLHttpRequest" or request.method not in ("POST", "DELETE"):
        return HttpResponseNotAllowed("Only post and delete ajax requests are allowed")
    try:
        data = json.loads(
            request.body.decode(request.encoding or settings.DEFAULT_CHARSET)
        )
        oldname = data.get("oldname", None)
        name = data.get("name", None)
        if not name and not oldname:
            return HttpResponse("Missing name attribute", status=400)
        elif request.method == "POST":
            if not request.user.has_perm("execute.add_scheduledtask"):
                return HttpResponse("Couldn't add or update scheduled task", status=401)
            obj, created = ScheduledTask.objects.using(request.database).get_or_create(
                name=oldname if oldname else name
            )
            if created:
                obj.user = request.user
            elif (
                not request.user.is_superuser
                and obj.user.username != request.user.username
            ):
                return HttpResponse(
                    "This task can only be updated by superusers and %s"
                    % obj.user.username,
                    status=401,
                )
            fld = data.get("email_failure", None)
            if fld is not None:
                obj.email_failure = fld
            fld = data.get("email_success", None)
            if fld is not None:
                obj.email_success = fld
            obj.tz = data.get("timezone", settings.TIME_ZONE)
            fld = data.get("data", None)
            if isinstance(fld, dict):
                obj.data = fld
            if oldname and name and oldname != name:
                # Rename the task
                obj.name = name
                ScheduledTask.objects.using(request.database).filter(
                    name=oldname
                ).delete()
            obj.save(using=request.database)
            scheduler.waitNextEvent(database=request.database)
            obj.adjustForTimezone(GridReport.getTimezoneOffset(request))
            return HttpResponse(
                content=(
                    obj.next_run.strftime("%Y-%m-%d %H:%M:%S") if obj.next_run else ""
                )
            )
        elif request.method == "DELETE":
            if not request.user.has_perm("execute.delete_scheduledtask"):
                return HttpResponse("Couldn't delete scheduled task", status=401)
            else:
                try:
                    obj = (
                        ScheduledTask.objects.using(request.database)
                        .get(name=oldname if oldname else name)
                        .delete()
                    )
                    return HttpResponse(content="OK")
                except ScheduledTask.DoesNotExist:
                    return HttpResponse("Couldn't delete scheduled task", status=400)
    except Exception as e:
        logger.error("Error updating scheduled task: %s" % e)
        return HttpResponseServerError("Error updating scheduled task")


@staff_member_required
@never_cache
def scenario_add(request):
    if request.method not in ("POST",):
        return HttpResponseNotAllowed("Only post requests are allowed")
    if not request.user.has_perm("auth.run_db"):
        return HttpResponse("No permission to add a scenario", status=401)
    try:
        error_code = updateScenarioCount(addition=True)
        if not error_code:
            return HttpResponse("Successfully added a new scenario")
        elif error_code == 2:
            return HttpResponse(
                "You have reached the maximum number of scenarios", status=401
            )
        elif error_code == 4:
            return HttpResponse("Invalid format of djangosettings.py file", status=401)
        else:
            return HttpResponse("An unknown error occured", status=400)
    except Exception as e:
        logger.error("Error adding scenario: %s" % e)
        return HttpResponse(e, status=400)


@staff_member_required
@never_cache
def scenario_delete(request):
    if request.method not in ("POST",):
        return HttpResponseNotAllowed("Only post requests are allowed")
    if not request.user.has_perm("auth.run_db"):
        return HttpResponse("No permission to delete a scenario", status=401)
    try:
        error_code = updateScenarioCount(addition=False)
        if not error_code:
            return HttpResponse("Successfully deleted a scenario")
        elif error_code == 1:
            return HttpResponse(
                "You have already reached the minimum number of scenarios", status=401
            )
        elif error_code == 3:
            return HttpResponse("Release your last scenario and try again", status=401)
        elif error_code == 4:
            return HttpResponse("Invalid format of djangosettings.py file", status=401)
        else:
            return HttpResponse("An unknown error occured", status=400)
    except Exception as e:
        logger.error("Error deleting scenario: %s" % e)
        return HttpResponse(e, status=400)


def exportWorkbook(request):
    # Create a workbook
    wb = Workbook(write_only=True)
    wb.properties.creator = "frepple %s" % __version__

    # Create a named style for the header row
    headerstyle = NamedStyle(name="headerstyle")
    headerstyle.fill = PatternFill(fill_type="solid", fgColor="70c4f4")
    wb.add_named_style(headerstyle)
    readlonlyheaderstyle = NamedStyle(name="readlonlyheaderstyle")
    readlonlyheaderstyle.fill = PatternFill(fill_type="solid", fgColor="d0ebfb")
    wb.add_named_style(readlonlyheaderstyle)

    # retrieve value of parameter excel_duration_in_days
    excel_duration_in_days = (
        Parameter.getValue("excel_duration_in_days", request.database, "false").lower()
        == "true"
    )

    # Loop over all selected entity types
    exportConfig = {"anonymous": request.POST.get("anonymous", False)}
    ok = False
    for entity_name in request.POST.getlist("entities"):
        try:
            # Initialize
            (app_label, model_label) = entity_name.split(".")
            model = apps.get_model(app_label, model_label)
            # Verify access rights
            permname = get_permission_codename("change", model._meta)
            if not request.user.has_perm("%s.%s" % (app_label, permname)):
                continue

            # Never export some special administrative models
            if model in EXCLUDE_FROM_BULK_OPERATIONS:
                continue

            # Create sheet
            ok = True
            ws = wb.create_sheet(title=force_str(model._meta.verbose_name))

            # Build a list of fields and properties
            fields = []
            modelfields = []
            header = []
            source = False
            lastmodified = False
            owner = False
            comment = None
            try:
                # The admin model of the class can define some fields to exclude from the export
                exclude = data_site._registry[model].exclude
            except Exception:
                exclude = None
            for i in model._meta.fields:
                if isinstance(i, AutoField) and i.primary_key:
                    # Don't export automatically generated primary keys.
                    # We rely on the natural for restoring the data.
                    continue
                elif i.name in ["lft", "rght", "lvl"]:
                    continue  # Skip some fields of HierarchyModel
                elif i.name == "source":
                    source = i  # Put the source field at the end
                elif i.name == "lastmodified":
                    lastmodified = i  # Put the last-modified field at the very end
                elif not (exclude and i.name in exclude):
                    fields.append(i.column)
                    modelfields.append(i)
                    cell = WriteOnlyCell(ws, value=force_str(i.verbose_name).title())
                    if i.editable:
                        cell.style = "headerstyle"
                        if isinstance(i, ForeignKey):
                            cell.comment = CellComment(
                                force_str(
                                    _("Values in this field must exist in the %s table")
                                    % force_str(i.remote_field.model._meta.verbose_name)
                                ),
                                "Author",
                            )
                        elif i.choices:
                            cell.comment = CellComment(
                                force_str(
                                    _("Accepted values are: %s")
                                    % ", ".join([c[0] for c in i.choices])
                                ),
                                "Author",
                            )
                    else:
                        cell.style = "readlonlyheaderstyle"
                        if not comment:
                            comment = CellComment(
                                force_str(_("Read only")),
                                "Author",
                                height=20,
                                width=80,
                            )
                        cell.comment = comment
                    header.append(cell)
                    if i.name == "owner":
                        owner = True
            if hasattr(model, "propertyFields"):
                if callable(model.propertyFields):
                    props = model.propertyFields(request)
                else:
                    props = model.propertyFields
                for i in props:
                    if i.export:
                        fields.append(i.name)
                        cell = WriteOnlyCell(
                            ws, value=force_str(i.verbose_name).title()
                        )
                        if i.editable:
                            cell.style = "headerstyle"
                            if isinstance(i, ForeignKey):
                                cell.comment = CellComment(
                                    force_str(
                                        _(
                                            "Values in this field must exist in the %s table"
                                        )
                                        % force_str(
                                            i.remote_field.model._meta.verbose_name
                                        )
                                    ),
                                    "Author",
                                )
                        elif i.choices:
                            cell.comment = CellComment(
                                force_str(
                                    _("Accepted values are: %s")
                                    % ", ".join([c[0] for c in i.choices])
                                ),
                                "Author",
                            )
                        else:
                            cell.style = "readlonlyheaderstyle"
                            if not comment:
                                comment = CellComment(
                                    force_str(_("Read only")),
                                    "Author",
                                    height=20,
                                    width=80,
                                )
                            cell.comment = comment
                        header.append(cell)
                        modelfields.append(i)
            if source:
                fields.append("source")
                cell = WriteOnlyCell(ws, value=force_str(_("source")).title())
                cell.style = "headerstyle"
                header.append(cell)
                modelfields.append(source)
            if lastmodified:
                fields.append("lastmodified")
                cell = WriteOnlyCell(ws, value=force_str(_("last modified")).title())
                cell.style = "readlonlyheaderstyle"
                if not comment:
                    comment = CellComment(
                        force_str(_("Read only")), "Author", height=20, width=80
                    )
                cell.comment = comment
                header.append(cell)
                modelfields.append(lastmodified)

            # Write a formatted header row
            ws.append(header)

            # Add an auto-filter to the table
            ws.auto_filter.ref = "A1:%s1048576" % get_column_letter(len(header))

            # Use the default manager
            if issubclass(model, HierarchyModel):
                model.rebuildHierarchy(database=request.database)
                query = (
                    model.objects.all().using(request.database).order_by("lvl", "pk")
                )
            elif owner:
                # First export records with empty owner field
                query = (
                    model.objects.all().using(request.database).order_by("-owner", "pk")
                )
            else:
                query = model.objects.all().using(request.database).order_by("pk")

            # Special annotation of the export query
            if hasattr(model, "export_objects"):
                query = model.export_objects(query, request)

            # Loop over all records
            for rec in query.values_list(*fields):
                cells = []
                fld = 0
                for f in rec:
                    cells.append(
                        _getCellValue(
                            f,
                            field=modelfields[fld],
                            exportConfig=exportConfig,
                            excel_duration_in_days=excel_duration_in_days,
                        )
                    )
                    fld += 1
                ws.append(cells)
        except Exception:
            pass  # Silently ignore the error and move on to the next entity.

    # Not a single entity to export
    if not ok:
        raise Exception(_("Nothing to export"))

    # Write the excel from memory to a string and then to a HTTP response
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.getvalue(),
    )
    response["Content-Disposition"] = 'attachment; filename="frepple.xlsx"'
    response["Cache-Control"] = "no-cache, no-store"
    return response


def importWorkbook(request):
    """
    This method reads a spreadsheet in Office Open XML format (typically with
    the extension .xlsx or .ods).
    Each entity has a tab in the spreadsheet, and the first row contains
    the fields names.
    """
    # Build a list of all contenttypes
    all_models = [
        (ct.model_class(), ct.pk)
        for ct in ContentType.objects.all()
        if ct.model_class()
    ]

    # retrieve value of parameter excel_duration_in_days
    excel_duration_in_days = (
        Parameter.getValue("excel_duration_in_days", request.database, "false").lower()
        == "true"
    )

    try:
        # Find all models in the workbook
        for filename, file in request.FILES.items():
            yield "<strong>" + force_str(
                _("Processing file")
            ) + " " + filename + "</strong><br>"
            if filename.endswith(".xls"):
                yield _(
                    "Files in the old .XLS excel format can't be read.<br>Please convert them to the new .XLSX format."
                )
                continue
            elif (
                file.content_type
                != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ):
                yield _("Unsupported file format.")
                continue
            wb = load_workbook(filename=file, data_only=True)
            models = []
            for ws_name in wb.sheetnames:
                # Find the model
                model = None
                contenttype_id = None
                for m, ct in all_models:
                    if matchesModelName(ws_name, m):
                        model = m
                        contenttype_id = ct
                        break
                if not model or model in EXCLUDE_FROM_BULK_OPERATIONS:
                    yield '<div class="alert alert-warning">' + force_str(
                        _("Ignoring data in worksheet: %s") % ws_name
                    ) + "</div>"
                elif not request.user.has_perm(
                    "%s.%s"
                    % (
                        model._meta.app_label,
                        get_permission_codename("add", model._meta),
                    )
                ):
                    # Check permissions
                    yield '<div class="alert alert-danger">' + force_str(
                        _("You don't permissions to add: %s") % ws_name
                    ) + "</div>"
                else:
                    deps = set([model])
                    GridReport.dependent_models(model, deps)
                    models.append((ws_name, model, contenttype_id, deps))

            # Sort the list of models, based on dependencies between models
            models = GridReport.sort_models(models)

            # Process all rows in each worksheet
            yield (
                '<div class="table-responsive">'
                '<table class="table table-sm" style="white-space: nowrap;"><tbody>'
            )
            for ws_name, model, contenttype_id, dependencies in models:
                with transaction.atomic(using=request.database):
                    yield '<tr><th colspan="5">%s %s<div class="recordcount"></div></th></tr>' % (
                        capfirst(_("worksheet")),
                        ws_name,
                    )
                    numerrors = 0
                    numwarnings = 0
                    firsterror = True
                    ws = wb[ws_name]
                    for error in parseExcelWorksheet(
                        model,
                        ws,
                        user=request.user,
                        database=request.database,
                        ping=True,
                        excel_duration_in_days=excel_duration_in_days,
                    ):
                        if error[0] == logging.DEBUG:
                            # Yield some result so we can detect disconnect clients and interrupt the upload
                            yield "<tr class='hidden' data-cnt='%s'>" % error[1]
                            continue
                        if firsterror and error[0] in (logging.ERROR, logging.WARNING):
                            yield '<tr><th class="sr-only">%s</th><th>%s</th><th>%s</th><th>%s</th><th>%s%s%s</th></tr>' % (
                                capfirst(_("worksheet")),
                                capfirst(_("row")),
                                capfirst(_("field")),
                                capfirst(_("value")),
                                capfirst(_("error")),
                                " / ",
                                capfirst(_("warning")),
                            )
                            firsterror = False
                        if error[0] == logging.ERROR:
                            yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                ws_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                capfirst(_("error")),
                                error[4],
                            )
                            numerrors += 1
                        elif error[1] == logging.WARNING:
                            yield '<tr><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s: %s</td></tr>' % (
                                ws_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                capfirst(_("warning")),
                                error[4],
                            )
                            numwarnings += 1
                        else:
                            yield '<tr class=%s><td class="sr-only">%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
                                "danger" if numerrors > 0 else "success",
                                ws_name,
                                error[1] if error[1] else "",
                                error[2] if error[2] else "",
                                error[3] if error[3] else "",
                                error[4],
                            )

                    # Excel duration might have changed after reading the parameter tab
                    if str(model._meta) == "common.parameter":
                        # retrieve value of parameter excel_duration_in_days
                        excel_duration_in_days = (
                            Parameter.getValue(
                                "excel_duration_in_days", request.database, "false"
                            ).lower()
                            == "true"
                        )
            yield "</tbody></table></div>"
            yield "<div><strong>%s</strong><br><br></div>" % _("Done")
    except GeneratorExit:
        logger.warning("Connection Aborted")
    except Exception as e:
        yield "Import aborted: %s" % e
        logger.error("Exception importing workbook: %s" % e)


@staff_member_required
@never_cache
def exports(request):
    if request.headers.get(
        "x-requested-with"
    ) != "XMLHttpRequest" or request.method not in ("POST", "DELETE"):
        return HttpResponseNotAllowed("Only post ajax requests are allowed")
    if not request.user.is_superuser:
        return HttpResponseForbidden(force_str(_("Only superusers can do this")))

    try:
        data = json.loads(
            request.body.decode(request.encoding or settings.DEFAULT_CHARSET)
        )
        errors = []
        curname = data.get("currentname", None)
        if data.get("delete", None):
            deleted = (
                DataExport.objects.using(request.database)
                .filter(name=data["delete"])
                .delete()
            )
            if deleted:
                # Remove the data file for security reasons and to save disk space
                try:
                    os.remove(
                        os.path.join(
                            get_databases()[request.database]["FILEUPLOADFOLDER"],
                            "export",
                            data["delete"],
                        )
                    )
                except FileNotFoundError:
                    pass
        elif data.get("name", None) in (None, ".xlsx", ".csv", ".csv.gz"):
            errors.append("Name can't be blank<br>")
        else:
            if not data["name"].endswith((".xlsx", ".csv", ".csv.gz")):
                errors.append("Export must end with .xlsx, .csv or .csv.gz<br>")
            elif (
                data["name"] != curname
                and DataExport.objects.using(request.database)
                .filter(name=data["name"])
                .exists()
            ):
                errors.append("Export with this name already exists<br>")
            elif os.sep in data["name"]:
                errors.append("Export names can't contain %s" % os.sep)
            else:
                t = data.get("type", None)
                if t == "sql":
                    s = data.get("sql", None)
                    if s:
                        if curname:
                            DataExport.objects.using(request.database).filter(
                                name=curname
                            ).update(name=data["name"], sql=s)
                        else:
                            DataExport(name=data["name"], sql=s).save(
                                using=request.database
                            )
                    else:
                        errors.append("Missing sql query<br>")
                elif t == "report":
                    r = data.get("report", None)
                    if not r or r not in (
                        "freppledb.output.views.resource.OverviewReport",
                        "freppledb.output.views.demand.OverviewReport",
                        "freppledb.output.views.buffer.OverviewReport",
                        "freppledb.output.views.operation.OverviewReport",
                        "freppledb.output.views.operation.DistributionReport",
                        "freppledb.output.views.operation.PurchaseReport",
                        "freppledb.forecast.views.OverviewReport",
                    ):
                        errors.append("Invalid report<br>")
                    else:
                        if curname:
                            DataExport.objects.using(request.database).filter(
                                name=curname
                            ).update(
                                name=data["name"],
                                report=r,
                                sql=None,
                                arguments={
                                    "buckets": data.get("bucket"),
                                    "horizontype": True,
                                    "horizonunit": data.get("horizonbucket", "month"),
                                    "horizonlength": int(data.get("horizon", 6)),
                                },
                            )
                        else:
                            DataExport(
                                name=data["name"],
                                report=r,
                                arguments={
                                    "buckets": data.get("bucket"),
                                    "horizontype": True,
                                    "horizonunit": data.get("horizonbucket", "month"),
                                    "horizonlength": int(data.get("horizon", 6)),
                                },
                            ).save(using=request.database)
                elif t == "customreport":
                    from freppledb.reportmanager.models import SQLReport

                    r = data.get("report", None)
                    if (
                        r
                        and SQLReport.objects.using(request.database)
                        .filter(id=r)
                        .exists()
                    ):
                        if curname:
                            DataExport.objects.using(request.database).filter(
                                name=curname
                            ).update(
                                name=data["name"],
                                report="freppledb.reportmanager.models.SQLReport.%s"
                                % r,
                                sql=None,
                            )
                        else:
                            DataExport(
                                name=data["name"],
                                report="freppledb.reportmanager.models.SQLReport.%s"
                                % r,
                            ).save(using=request.database)
                    else:
                        errors.append("Invalid custom report<br>")
                else:
                    errors.append("Unknown export type<br>")
        if errors:
            return HttpResponseServerError(content="\n".join(errors))
        else:
            return HttpResponse(content="OK")
    except Exception as e:
        logger.error("Error updating export: %s" % e)
        return HttpResponseServerError("Error updating export")
